"""Build the single canonical perk list.

Source of truth: Reference xlsx "Unabridged List" sheet, filtered to
Status != "Excluded". Multi-perk rows (Name cell containing one or
more "\n" separators) are split into one entry per perk; all split
rows inherit the parent row's Constellation, Jumpdoc and Status.

Each entry is then enriched by joining against:
  - perks_catalog.json    -> cost + description
  - obtained_perks.json   -> acquired_chapter (and the derived `acquired` bool)

For each join the script attempts in order:
  1. exact (name, source) match
  2. name-only match (cosmetic-tolerant: lowercase, fold curly punct,
     non-word chars -> space; same algorithm as parse_reference._normalize)
  3. no match -> null

The constellation written is the constellation from the Unabridged
List row (already canonical). The hand-curated overrides file
(perk_constellation_overrides.json) is consulted as a tiebreaker:
when a catalog/obtained match disagrees with the Unabridged List, the
override picks the constellation - but since constellation is sourced
from the Unabridged List directly, the override only acts as a
sanity check at build time and is otherwise informational.

Output:
  data/derived/canonical_perks.json   (validated against
  data/derived/_schemas/canonical_perks.schema.json)
"""

from __future__ import annotations

import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from _common import write_validated_json

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Brocktons_Celestial_Forge_Reference.xlsx"
PERKS_CATALOG_JSON = ROOT / "data" / "derived" / "perks_catalog.json"
OBTAINED_PERKS_JSON = ROOT / "data" / "derived" / "obtained_perks.json"
OVERRIDES_JSON = ROOT / "data" / "manual" / "perk_constellation_overrides.json"
OUT = ROOT / "data" / "derived" / "canonical_perks.json"

CONSTELLATIONS = {
    "Toolkits", "Knowledge", "Vehicles", "Time", "Crafting",
    "Clothing", "Magic", "Quality", "Size",
    "Resources and Durability", "Magitech", "Alchemy",
    "Capstone", "Personal Reality", "Felyne Perks",
}
ALLOWED_STATUSES = {
    "Obtained", "Available", "Partial",
    "Locked", "Unknown", "Repeatable",
}


def _normalize(s: str | None) -> str:
    """Cosmetic-tolerant normalization for fuzzy matching.

    Lowercase, fold curly quotes to ASCII, replace any non-word char
    (other than apostrophe) with space, collapse whitespace. Mirrors
    the helper in scripts/parse_reference.py so name matching is
    consistent across the pipeline.
    """
    if not s:
        return ""
    out = s.lower()
    for a, b in [("’", "'"), ("‘", "'")]:
        out = out.replace(a, b)
    out = re.sub(r"[^\w\s']", " ", out)
    out = re.sub(r"\s+", " ", out).strip()
    return out


@dataclass
class CanonicalPerk:
    constellation: str
    name: str
    jump: str | None
    status: str
    cost: int | None
    description: str | None
    acquired: bool
    acquired_chapter: str | None
    match_quality: str   # "exact" | "name_only" | "no_match"


# ---------- index builders --------------------------------------------------


def _build_catalog_indexes(catalog_path: Path) -> tuple[
    dict[tuple[str, str], dict[str, Any]],
    dict[str, dict[str, Any]],
]:
    """Return ((name_norm, source_norm) -> perk dict, name_norm -> perk dict).

    First-seen wins for the name-only index so behaviour is deterministic
    given the catalog's natural order.
    """
    by_exact: dict[tuple[str, str], dict[str, Any]] = {}
    by_name: dict[str, dict[str, Any]] = {}
    if not catalog_path.exists():
        return by_exact, by_name
    catalog = json.loads(catalog_path.read_text())
    for p in catalog.get("perks", []):
        nn = _normalize(p.get("name"))
        sn = _normalize(p.get("source"))
        if not nn:
            continue
        by_exact.setdefault((nn, sn), p)
        by_name.setdefault(nn, p)
    return by_exact, by_name


def _build_obtained_indexes(obtained_path: Path) -> tuple[
    dict[tuple[str, str], dict[str, Any]],
    dict[str, dict[str, Any]],
]:
    """Return ((name_norm, source_norm) -> obtained dict, name_norm -> obtained dict).

    First-seen wins (= earliest acquisition for repeatable perks).
    """
    by_exact: dict[tuple[str, str], dict[str, Any]] = {}
    by_name: dict[str, dict[str, Any]] = {}
    if not obtained_path.exists():
        return by_exact, by_name
    obtained = json.loads(obtained_path.read_text())
    for p in obtained.get("perks", []):
        nn = _normalize(p.get("perk_name"))
        sn = _normalize(p.get("jump"))
        if not nn:
            continue
        by_exact.setdefault((nn, sn), p)
        by_name.setdefault(nn, p)
    return by_exact, by_name


def _load_overrides(overrides_path: Path) -> dict[tuple[str, str], str]:
    if not overrides_path.exists():
        return {}
    data = json.loads(overrides_path.read_text())
    out: dict[tuple[str, str], str] = {}
    for key, val in data.get("overrides", {}).items():
        if "|" in key:
            name, source = key.split("|", 1)
        else:
            name, source = key, ""
        out[(_normalize(name), _normalize(source))] = val["constellation"]
    return out


# ---------- join helpers ----------------------------------------------------


def _lookup(
    name: str,
    source: str | None,
    by_exact: dict[tuple[str, str], dict[str, Any]],
    by_name: dict[str, dict[str, Any]],
) -> tuple[dict[str, Any] | None, str]:
    """Return (matched dict or None, match_quality)."""
    nn = _normalize(name)
    sn = _normalize(source)
    if not nn:
        return None, "no_match"
    if (nn, sn) in by_exact:
        return by_exact[(nn, sn)], "exact"
    if nn in by_name:
        return by_name[nn], "name_only"
    return None, "no_match"


_QUALITY_RANK = {"exact": 2, "name_only": 1, "no_match": 0}


def _best(*qualities: str) -> str:
    return max(qualities, key=lambda q: _QUALITY_RANK[q])


# ---------- main build ------------------------------------------------------


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", "").strip()


def build_canonical_perks() -> list[CanonicalPerk]:
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Unabridged List"]

    cat_exact, cat_name = _build_catalog_indexes(PERKS_CATALOG_JSON)
    obt_exact, obt_name = _build_obtained_indexes(OBTAINED_PERKS_JSON)
    overrides = _load_overrides(OVERRIDES_JSON)

    perks: list[CanonicalPerk] = []
    surprises: list[str] = []

    for r in range(2, ws.max_row + 1):
        constellation = _norm_text(ws.cell(r, 1).value)
        names_raw = ws.cell(r, 2).value
        jump = _norm_text(ws.cell(r, 3).value) or None
        status = _norm_text(ws.cell(r, 4).value)

        if not names_raw:
            continue
        if status == "Excluded":
            continue
        if status not in ALLOWED_STATUSES:
            surprises.append(f"row {r}: unknown status {status!r} (skipped)")
            continue
        if constellation not in CONSTELLATIONS:
            # Pick a sensible default and report; we don't want to crash
            # because of one stray label.
            surprises.append(
                f"row {r}: unknown constellation {constellation!r}; "
                f"defaulting to 'Knowledge'"
            )
            constellation = "Knowledge"

        for raw_name in str(names_raw).split("\n"):
            name = raw_name.strip()
            if not name:
                continue

            cat_hit, cat_q = _lookup(name, jump, cat_exact, cat_name)
            obt_hit, obt_q = _lookup(name, jump, obt_exact, obt_name)

            cost = cat_hit["cost"] if cat_hit else None
            description = cat_hit["description"] if cat_hit else None
            acquired_chapter = obt_hit["chapter_num"] if obt_hit else None
            acquired = acquired_chapter is not None

            perks.append(
                CanonicalPerk(
                    constellation=constellation,
                    name=name,
                    jump=jump,
                    status=status,
                    cost=cost,
                    description=description,
                    acquired=acquired,
                    acquired_chapter=acquired_chapter,
                    match_quality=_best(cat_q, obt_q),
                )
            )

    # Override sanity check: surface (don't enforce) any disagreement.
    for p in perks:
        key = (_normalize(p.name), _normalize(p.jump))
        if key in overrides and overrides[key] != p.constellation:
            surprises.append(
                f"override mismatch: {p.name!r} ({p.jump!r}) "
                f"override={overrides[key]!r} vs Unabridged={p.constellation!r}"
            )

    if surprises:
        print("structural surprises:")
        for s in surprises:
            print(f"  - {s}")

    return perks


def main() -> None:
    perks = build_canonical_perks()
    status_dist = dict(Counter(p.status for p in perks))
    acquired_count = sum(1 for p in perks if p.acquired)
    match_dist = Counter(p.match_quality for p in perks)

    payload = {
        "_source": (
            "data/raw/Brocktons_Celestial_Forge_Reference.xlsx#Unabridged List "
            "(Status != 'Excluded'); enriched from data/derived/perks_catalog.json "
            "and data/derived/obtained_perks.json"
        ),
        "_count": len(perks),
        "_status_distribution": status_dist,
        "_acquired_count": acquired_count,
        "_note": (
            "Source of truth is the Reference xlsx Unabridged List sheet, filtered "
            "to Status != 'Excluded'. Multi-perk rows (Name cell with newline "
            "separators) are split into one entry per perk. cost/description are "
            "joined from perks_catalog.json (exact (name, source) -> name-only -> "
            "null). acquired_chapter is joined from obtained_perks.json the same "
            "way; acquired is true iff acquired_chapter is set. match_quality is "
            "the best (most specific) join result across the cost and acquired "
            "lookups: 'exact' beats 'name_only' beats 'no_match'. Constellation "
            "is taken directly from the Unabridged List row."
        ),
        "perks": [asdict(p) for p in perks],
    }

    write_validated_json(OUT, payload, "canonical_perks")

    pct = (100.0 * acquired_count / len(perks)) if perks else 0.0
    unmatched = match_dist.get("no_match", 0)
    print(f"wrote {OUT.relative_to(ROOT)}: {len(perks)} perks")
    print(f"  status distribution: {status_dist}")
    print(f"  acquired: {acquired_count}/{len(perks)} ({pct:.1f}%)")
    print(f"  match quality: {dict(match_dist)}")
    print(f"  unmatched (no catalog AND no obtained hit): {unmatched}")


if __name__ == "__main__":
    main()
