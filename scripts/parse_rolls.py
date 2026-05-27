"""Parse the rolls list xlsx into rolls.json and perks_catalog.json.

Source:  data/raw/Brocktons_Celestial_Forge_Rolls_List.xlsx
         (sonicyoash's curated sheet linked from threadmark info post #12)

Sheets used:
  - "List of Rolls & Perk Order"  -> data/derived/rolls.json
  - "Complete List of Perks"      -> data/derived/perks_catalog.json

Coverage caveat: the curator stopped maintaining the rolls list at
chapter 75; the story continues to ~ch 120. perks_catalog covers all
known constellation perks regardless of chapter.
"""

from __future__ import annotations

import re
from dataclasses import asdict, dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from _common import write_validated_json
from perk_name_resolver import parse_cost_text

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Brocktons_Celestial_Forge_Rolls_List.xlsx"
OUT_ROLLS = ROOT / "data" / "derived" / "rolls.json"
OUT_PERKS = ROOT / "data" / "derived" / "perks_catalog.json"


# ---------- rolls -----------------------------------------------------------


@dataclass
class Perk:
    name: str
    source: str           # jump source, e.g. "Highschool of the Dead"
    cost: int             # effective CP cost (non-CP units pre-multiplied x100); 0 for Free
    cost_text: str        # original cost text from the roll line, e.g. "200-100", "Free", "3 Customization Points"
    cost_unit: str | None # unit string when not CP/WP/blank, else None
    free: bool


@dataclass
class Roll:
    roll_number: int | None     # None for Trigger Event and annotations
    chapter_num: str            # raw prefix from sheet col A, e.g. "1", "3.1"
    kind: str                   # "trigger" | "roll" | "miss" | "annotation"
    banked_before: int | None   # CP available going into the roll
    banked_after: int | None    # CP banked after
    constellation: str | None   # rolled constellation; None for annotations
    constellation_revealed: bool
    perks: list[Perk] = field(default_factory=list)
    raw: str = ""


# main roll line: "Roll N (X): Constellation → Result → (Y)"
# also matches "Roll Number N" and "Trigger Event". The curator
# occasionally uses "=" instead of "→" for the closing separator.
_ARROW = r"(?:→|->|=)"
_HEAD_RE = re.compile(
    r"^(?P<head>Trigger Event|Roll(?:\s+Number)?\s+(?P<num>\d+))"
    r"\s*\((?P<before>\d+)\)\s*:\s*"
    r"(?P<const>.+?)"
    rf"\s*{_ARROW}\s*"
    r"(?P<result>.+?)"
    rf"\s*{_ARROW}\s*"
    r"\((?P<after>\d+)\)\s*$"
)
_TRIGGER_WORKSHOP_SOURCE = (
    "Trigger Event (100): Toolkits Constellation Revealed → "
    "Workshop (Personal Reality Supplement) (100-100) + "
    "Access Key (Personal Reality Supplement) (Free) + "
    "Entrance Hall (Personal Reality Supplement) (Free) → (0)"
)
_TRIGGER_WORKSHOP_DISCOUNTED = (
    "Trigger Event (0): Toolkits Constellation Revealed → "
    "Workshop (Personal Reality Supplement) (0-0) + "
    "Access Key (Personal Reality Supplement) (Free) + "
    "Entrance Hall (Personal Reality Supplement) (Free) → (0)"
)
_CURATED_COVERAGE_NOTE = (
    "chapters 1-75 (curator stopped maintaining beyond ch 75) | "
    "Curated: rolls 61/207/319 promoted from annotation→roll/miss; "
    "roll 98 constellation Tollkits→Toolkits (raw preserved)"
)
_CH351_MALFORMED_ROLL_LINES = {
    "Roll ? (<=300): Crafting → Weaponsmith (Light of Terra DLC 5 A Sky Filled With Steel - Warhammer 40,000) (<=300-300) → (<=0)",
    "Roll ? (<=400): Knowledge → I Am Iron Man (Marvel Cinematic Universe) (<=400-400) → (<=0)",
    "Roll 207 (<=200): ??? → ??? (<=200-<=0) → (200)",
    "8 Rolls occured during this interlude with 6 Misses and 2 Perks acquired (100 Banked + 800 - 700 Points worth of acquired Perks = 200)",
}
_CH351_OVERLAY_ROLL_LINES = [
    "Roll 200 (200): ??? → Miss → (200)",
    "Roll 201 (300): Crafting → Weaponsmith (Light of Terra DLC 5 A Sky Filled With Steel - Warhammer 40,000) (300-300) → (0)",
    "Roll 202 (100): ??? → Miss → (100)",
    "Roll 203 (200): ??? → Miss → (200)",
    "Roll 204 (300): ??? → Miss → (300)",
    "Roll 205 (400): Knowledge → I Am Iron Man (Marvel Cinematic Universe) (400-400) → (0)",
    "Roll 206 (100): ??? → Miss → (100)",
    "Roll 207 (200): ??? → Miss → (200)",
]
# perk piece in result. The source is optional (some Personal Reality
# perks don't list one). Cost forms:
#   - "X-Y" (banked_before-banked_after style; Y is the actual CP cost)
#   - "N Customization Points" (non-CP unit; the ×100 rule applies)
#   - any "Free..." variant ("Free", "Free?", "Free Soldier", ...).
_PERK_RE = re.compile(
    r"^(?P<name>.+?)\s*"
    r"(?:\((?P<source>[^()]*(?:\([^()]*\)[^()]*)*)\)\s*)?"
    r"\((?P<cost>\d+-\d+|\d+\s+[A-Za-z][A-Za-z\s]*|Free[^)]*)\)\s*$"
)


def _parse_constellation(raw: str) -> tuple[str, bool]:
    revealed = False
    s = raw.strip()
    if s.endswith("Constellation Revealed"):
        s = s[: -len("Constellation Revealed")].strip()
        revealed = True
    elif s.endswith("Constellation"):
        s = s[: -len("Constellation")].strip()
    return s, revealed


def _parse_perk_chain(result: str) -> list[Perk]:
    """A perk-acquisition result is a "+"-joined chain of perk pieces."""
    pieces = [p.strip() for p in result.split(" + ")]
    perks: list[Perk] = []
    for piece in pieces:
        m = _PERK_RE.match(piece)
        if not m:
            # Best-effort: store the whole piece as the name with empty source.
            perks.append(Perk(
                name=piece, source="", cost=0,
                cost_text="", cost_unit=None, free=False,
            ))
            continue
        cost_str = m.group("cost").strip()
        if cost_str.lower().startswith("free"):
            cost = 0
            cost_unit = None
            free = True
        elif "-" in cost_str:
            # "X-Y" form: Y is the actual CP cost (pure CP).
            cost = int(cost_str.split("-", 1)[1])
            cost_unit = None
            free = False
        else:
            # Non-CP unit form (e.g. "3 Customization Points"). Share
            # the same ×100 intercept logic as the catalog and obtained-
            # perks parsers via perk_name_resolver.parse_cost_text.
            cost, free, cost_unit = parse_cost_text(cost_str)
        source = m.group("source") or ""
        perks.append(
            Perk(
                name=m.group("name").strip(),
                source=source.strip(),
                cost=cost,
                cost_text=cost_str,
                cost_unit=cost_unit,
                free=free,
            )
        )
    return perks


def _parse_roll_line(line: str, chapter_num: str) -> Roll:
    raw = line.strip()
    if raw == _TRIGGER_WORKSHOP_SOURCE:
        raw = _TRIGGER_WORKSHOP_DISCOUNTED
    if raw == (
        "Roll 61 (400): Innate Talent: Alchemist (Overlord) (400-200) + "
        "Alchemist's Laboratory (Overlord) (Free) → (200)"
    ):
        return Roll(
            roll_number=61,
            chapter_num=chapter_num,
            kind="roll",
            banked_before=400,
            banked_after=200,
            constellation="Alchemy",
            constellation_revealed=False,
            perks=_parse_perk_chain(
                "Innate Talent: Alchemist (Overlord) (400-200) + "
                "Alchemist's Laboratory (Overlord) (Free)"
            ),
            raw=raw,
        )
    if raw == "Roll 207 (<=200): ??? → ??? (<=200-<=0) → (200)":
        return Roll(
            roll_number=207,
            chapter_num=chapter_num,
            kind="miss",
            banked_before=200,
            banked_after=200,
            constellation=None,
            constellation_revealed=False,
            raw=raw,
        )
    if raw == "Roll 319 (200): Vehicles → Miss (200)":
        return Roll(
            roll_number=319,
            chapter_num=chapter_num,
            kind="miss",
            banked_before=200,
            banked_after=200,
            constellation="Vehicles",
            constellation_revealed=False,
            raw=raw,
        )
    m = _HEAD_RE.match(raw)
    if not m:
        return Roll(
            roll_number=None,
            chapter_num=chapter_num,
            kind="annotation",
            banked_before=None,
            banked_after=None,
            constellation=None,
            constellation_revealed=False,
            raw=raw,
        )

    head = m.group("head")
    num_str = m.group("num")
    constellation, revealed = _parse_constellation(m.group("const"))
    if constellation == "Tollkits":
        constellation = "Toolkits"
    result = m.group("result").strip()
    is_miss = result.lower() == "miss"

    return Roll(
        roll_number=int(num_str) if num_str else None,
        chapter_num=chapter_num,
        kind="trigger" if head == "Trigger Event" else ("miss" if is_miss else "roll"),
        banked_before=int(m.group("before")),
        banked_after=int(m.group("after")),
        constellation=constellation if constellation and constellation != "???" else None,
        constellation_revealed=revealed,
        perks=[] if is_miss else _parse_perk_chain(result),
        raw=raw,
    )


def parse_rolls(wb) -> list[Roll]:
    ws = wb["List of Rolls & Perk Order"]
    rolls: list[Roll] = []
    emitted_ch351_overlay = False
    for r in range(2, ws.max_row + 1):
        chapter = ws.cell(r, 1).value
        body = ws.cell(r, 2).value
        if not body:
            continue
        chapter_num = str(chapter).split(" ", 1)[0] if chapter else ""
        for line in str(body).splitlines():
            line = line.strip()
            if not line:
                continue
            if chapter_num == "35.1" and line in _CH351_MALFORMED_ROLL_LINES:
                if not emitted_ch351_overlay:
                    rolls.extend(
                        _parse_roll_line(overlay_line, chapter_num)
                        for overlay_line in _CH351_OVERLAY_ROLL_LINES
                    )
                    emitted_ch351_overlay = True
                continue
            rolls.append(_parse_roll_line(line, chapter_num))
    return rolls


# ---------- perks catalog ---------------------------------------------------


@dataclass
class CatalogPerk:
    constellation: str
    name: str
    source: str
    cost: int               # effective CP cost (non-CP units pre-multiplied x100)
    cost_text: str          # raw "100 CP", "Free", "3 Customization Points"
    cost_unit: str | None   # unit string when not CP / blank / pure number, else None
    repeatable: bool
    description: str


def _parse_cost_text(cost_text: str) -> tuple[int, str | None]:
    """Parse a catalog cost-text. Thin wrapper around the shared
    ``perk_name_resolver.parse_cost_text`` that drops the ``is_free``
    return value (the catalog separates Free entries by other means)."""
    cp, _free, unit = parse_cost_text(cost_text)
    return cp, unit


# A cell in the Complete List of Perks looks like:
#   "Workshop (Personal Reality) (100 CP) - Repeatable\nEach purchase of this..."
# Source is optional (Personal Reality items often omit it). Cost may
# have any unit suffix (CP / WP / Customization Points / none) or be
# any "Free..." variant.
_CATALOG_HEAD_RE = re.compile(
    r"^(?P<name>.+?)\s*"
    r"(?:\((?P<source>[^()]*(?:\([^()]*\)[^()]*)*)\)\s*)?"
    r"\((?P<cost>\d+(?:\s*[A-Za-z]+(?:\s+[A-Za-z]+)*)?|Free[^)]*)\)"
    r"(?:\s*-\s*(?P<flag>Repeatable))?"
    r"\s*\n?(?P<desc>.*)$",
    re.DOTALL,
)


def parse_perks_catalog(wb) -> list[CatalogPerk]:
    ws = wb["Complete List of Perks"]
    # Header row 1 has constellation labels in even columns (2, 4, 6, ...).
    headers: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and "Constellation" in str(v):
            headers[c] = str(v).replace(" Constellation", "").strip()

    perks: list[CatalogPerk] = []
    for col, constellation in headers.items():
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, col).value
            if not cell:
                continue
            text = str(cell).strip()
            m = _CATALOG_HEAD_RE.match(text)
            if not m:
                # Last column "Notes" has legend rows like "This means that this Perk has been Purchased"
                continue
            cost_text = m.group("cost").strip()
            cost, cost_unit = _parse_cost_text(cost_text)
            source = m.group("source") or ""
            perks.append(
                CatalogPerk(
                    constellation=constellation,
                    name=m.group("name").strip(),
                    source=source.strip(),
                    cost=cost,
                    cost_text=cost_text,
                    cost_unit=cost_unit,
                    repeatable=m.group("flag") == "Repeatable",
                    description=m.group("desc").strip(),
                )
            )
    return perks


# ---------- main ------------------------------------------------------------


def main() -> None:
    wb = load_workbook(SRC, data_only=True)

    rolls = parse_rolls(wb)
    write_validated_json(
        OUT_ROLLS,
        {
            "_source": "data/raw/Brocktons_Celestial_Forge_Rolls_List.xlsx#List of Rolls & Perk Order",
            "_count": len(rolls),
            "_coverage": _CURATED_COVERAGE_NOTE,
            "_kinds": "roll | miss | trigger | annotation",
            "rolls": [asdict(r) for r in rolls],
        },
        "rolls",
    )

    perks = parse_perks_catalog(wb)
    write_validated_json(
        OUT_PERKS,
        {
            "_source": "data/raw/Brocktons_Celestial_Forge_Rolls_List.xlsx#Complete List of Perks",
            "_count": len(perks),
            "perks": [asdict(p) for p in perks],
        },
        "perks_catalog",
    )

    # Summary
    by_kind: dict[str, int] = {}
    for r in rolls:
        by_kind[r.kind] = by_kind.get(r.kind, 0) + 1
    print(f"wrote {OUT_ROLLS.relative_to(ROOT)}: {len(rolls)} entries  {by_kind}")

    by_const: dict[str, int] = {}
    for p in perks:
        by_const[p.constellation] = by_const.get(p.constellation, 0) + 1
    print(f"wrote {OUT_PERKS.relative_to(ROOT)}: {len(perks)} perks across {len(by_const)} constellations")


if __name__ == "__main__":
    main()
