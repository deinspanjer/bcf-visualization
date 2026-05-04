"""Parse the second curator xlsx into derived JSON.

Source:  data/raw/Brocktons_Celestial_Forge_Reference.xlsx

Outputs:
  data/derived/obtained_perks.json   - full-story chronological acquisitions
  data/derived/timeline.json         - in-world timeline (replaces the
                                       previous Whamodyne-only version;
                                       this sheet also includes Whamodyne
                                       data plus pre-story dates from
                                       the author)

Sheets used:
  - "Obtained Perks"      -> obtained_perks.json
  - "Timeline of Events"  -> timeline.json
"""

from __future__ import annotations

import datetime as dt
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

from _common import write_validated_json

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Brocktons_Celestial_Forge_Reference.xlsx"
PERKS_CATALOG_JSON = ROOT / "data" / "derived" / "perks_catalog.json"
OVERRIDES_JSON = ROOT / "data" / "manual" / "perk_constellation_overrides.json"
OUT_PERKS = ROOT / "data" / "derived" / "obtained_perks.json"
OUT_TIMELINE = ROOT / "data" / "derived" / "timeline.json"


# ---------- obtained perks --------------------------------------------------


@dataclass
class ObtainedPerk:
    epub_sequence: int          # the Reference sheet's column A
    chapter_num: str            # parsed prefix from the chapter title (e.g. "97" or "3.1")
    chapter_full_title: str     # original title text (e.g. "97 Confrontations - Preamble Carol")
    perk_name: str
    classification: str | None  # e.g. "Facility", "System", "Fiat, System"; None if blank
    jump: str | None            # source jump (e.g. "Personal Reality"); None if blank
    cost: int                   # numeric CP cost (0 for free variants and Customization-Points perks)
    cost_text: str              # original cost text, e.g. "100", "Free", "Free with Alchemist"
    free: bool                  # any "Free..." variant counts
    perk_text: str
    constellation: str | None   # joined from catalog + Unabridged List; None if no match


_CONSTELLATIONS = {
    "Toolkits", "Knowledge", "Vehicles", "Time", "Crafting",
    "Clothing", "Magic", "Quality", "Size",
    "Resources and Durability", "Magitech", "Alchemy",
    "Capstone", "Personal Reality", "Felyne Perks",
}


def _normalize(s: str) -> str:
    """Cosmetic-tolerant normalization for fuzzy matching.

    Lowercase, fold curly quotes to ASCII, replace any non-word char
    (other than apostrophe) with space, collapse whitespace. Treats
    "-", ":", "—" as separators rather than preserved characters so
    that "Valuable Memories -the X" and "Valuable Memories: the X"
    both reduce to "valuable memories the x".
    """
    if not s:
        return ""
    out = s.lower()
    for a, b in [("’", "'"), ("‘", "'")]:
        out = out.replace(a, b)
    out = re.sub(r"[^\w\s']", " ", out)
    out = re.sub(r"\s+", " ", out).strip()
    return out


# Known generic catalog entries that the obtained list expands into
# many specific instances the catalog doesn't enumerate. Maps a regex
# matching the obtained perk name -> constellation.
_GENERIC_PATTERNS: list[tuple[re.Pattern, str]] = [
    # Percy Jackson: catalog has "Minor blessings" (plural, generic);
    # obtained has "Minor Blessing <God> - <Domain>" per blessing acquired.
    (re.compile(r"^minor\s+blessing\b", re.I), "Quality"),
]


# A few constellation names are also used as in-perk prefixes ("Time:",
# "Resources:") in the obtained list. Map to the canonical constellation.
_CONSTELLATION_NAME_PREFIXES: dict[str, str] = {
    "time": "Time",
    "resources": "Resources and Durability",
    "knowledge": "Knowledge",
    "vehicles": "Vehicles",
    "crafting": "Crafting",
    "clothing": "Clothing",
    "magic": "Magic",
    "quality": "Quality",
    "size": "Size",
    "magitech": "Magitech",
    "alchemy": "Alchemy",
    "capstone": "Capstone",
    "toolkits": "Toolkits",
}


_PREFIX_SEPARATORS = (":", " - ", " – ", " — ")


def _prefix_candidates(name: str) -> list[str]:
    """Yield progressively shorter prefixes of `name`, useful for
    matching "Workshop: Metalworking" -> "Workshop" or
    "Unnatural Skill: Smith" -> "Unnatural Skill".
    """
    cands = []
    seen = {name}
    for sep in _PREFIX_SEPARATORS:
        if sep in name:
            prefix = name.split(sep, 1)[0].strip()
            if prefix and prefix not in seen:
                cands.append(prefix)
                seen.add(prefix)
    return cands


def _build_constellation_index(wb) -> dict[str, dict[str, str]]:
    """Build a multi-key lookup from perk identifiers -> constellation.

    Sources, in priority order:
      1. perks_catalog.json (Complete List of Perks sheet)
      2. Reference xlsx Unabridged List sheet

    Returned dict has three sub-dicts: by `(name, source)` exact,
    by `(name, source)` normalized, and by `name` normalized only.
    """
    by_name_source: dict[tuple[str, str], str] = {}
    by_name_source_norm: dict[tuple[str, str], str] = {}
    by_name_norm: dict[str, str] = {}

    def add(constellation, name, source):
        if not constellation or not name:
            return
        c = constellation.strip()
        # Trim "Constellation" suffix if present
        if c.endswith(" Constellation"):
            c = c[: -len(" Constellation")]
        if c not in _CONSTELLATIONS:
            return
        n = name.strip()
        s = (source or "").strip()
        by_name_source.setdefault((n, s), c)
        nn = _normalize(n)
        ns = _normalize(s)
        by_name_source_norm.setdefault((nn, ns), c)
        by_name_norm.setdefault(nn, c)

    # 1. Existing catalog (already derived JSON)
    if PERKS_CATALOG_JSON.exists():
        catalog = json.loads(PERKS_CATALOG_JSON.read_text())
        for p in catalog.get("perks", []):
            add(p.get("constellation"), p.get("name"), p.get("source"))

    # 2. Reference xlsx Unabridged List (multi-perk rows split by newline)
    ws = wb["Unabridged List"]
    for r in range(2, ws.max_row + 1):
        constellation = ws.cell(r, 1).value
        names_raw = ws.cell(r, 2).value
        jump = ws.cell(r, 3).value
        if not names_raw:
            continue
        for name in str(names_raw).split("\n"):
            add(constellation, name, jump)

    return {
        "exact": by_name_source,
        "norm": by_name_source_norm,
        "name_only": by_name_norm,
    }


def _load_overrides() -> dict[tuple[str, str], str]:
    """Hand-curated (name, source) -> constellation mapping for perks
    that neither the catalog nor the Unabridged List nor any of the
    automated rules can classify. Each entry has a written reason.
    """
    if not OVERRIDES_JSON.exists():
        return {}
    data = json.loads(OVERRIDES_JSON.read_text())
    out: dict[tuple[str, str], str] = {}
    for key, val in data.get("overrides", {}).items():
        if "|" in key:
            name, source = key.split("|", 1)
        else:
            name, source = key, ""
        out[(name, source)] = val["constellation"]
    return out


def _classify(idx: dict[str, dict[str, str]], name: str, source: str | None,
              overrides: dict[tuple[str, str], str]) -> str | None:
    """Return the constellation for (name, source), trying progressively
    more permissive matches.
    """
    src = source or ""
    # 0. Manual overrides take precedence (curated perk-by-perk).
    if (name, src) in overrides:
        return overrides[(name, src)]
    # 1. Exact (name, source)
    if (name, src) in idx["exact"]:
        return idx["exact"][(name, src)]
    # 2. Normalized (name, source)
    nn = _normalize(name)
    ns = _normalize(src)
    if (nn, ns) in idx["norm"]:
        return idx["norm"][(nn, ns)]
    # 3. Normalized name (any source)
    if nn in idx["name_only"]:
        return idx["name_only"][nn]
    # 4. Prefix split fallbacks
    for prefix in _prefix_candidates(name):
        pn = _normalize(prefix)
        # 4a: prefix is a constellation name itself ("Time: X", "Resources: X")
        if pn in _CONSTELLATION_NAME_PREFIXES:
            return _CONSTELLATION_NAME_PREFIXES[pn]
        # 4b: prefix is a known catalog entry ("Workshop: X" -> "Workshop")
        if pn in idx["name_only"]:
            return idx["name_only"][pn]
    # 5. Generic-pattern fallbacks for catalogs that don't enumerate
    #    every specific instance (Percy Jackson "Minor blessings", etc.)
    for pattern, const in _GENERIC_PATTERNS:
        if pattern.search(name):
            return const
    # 6. Jump-as-constellation: Personal Reality jump perks default to
    #    the Personal Reality constellation (115 of 118 PR-jump catalog
    #    entries are Personal Reality; the 3 exceptions hit step 1-3 above).
    if src in _CONSTELLATIONS:
        return src
    return None


_CHAPTER_PREFIX_RE = re.compile(r"^(\d+(?:\.\d+)?)")


def _norm(value) -> str:
    return str(value).replace("\r", "").strip() if value is not None else ""


def _parse_cost(text: str) -> tuple[int, bool]:
    s = text.strip()
    low = s.lower()
    if low.startswith("free"):
        return 0, True
    # Numeric CP, possibly with a unit suffix like "2 Customization Points"
    m = re.match(r"^(\d+)", s)
    if m:
        return int(m.group(1)), False
    return 0, False


def parse_obtained_perks(wb, constellation_idx, overrides) -> list[ObtainedPerk]:
    ws = wb["Obtained Perks"]
    perks: list[ObtainedPerk] = []
    for r in range(2, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        if seq is None:
            continue
        title_raw = _norm(ws.cell(r, 2).value)
        prefix = _CHAPTER_PREFIX_RE.match(title_raw)
        chapter_num = prefix.group(1) if prefix else "0"
        cost_text = _norm(ws.cell(r, 6).value)
        cost, free = _parse_cost(cost_text)
        classification = _norm(ws.cell(r, 4).value) or None
        jump = _norm(ws.cell(r, 5).value) or None
        name = _norm(ws.cell(r, 3).value)
        perks.append(
            ObtainedPerk(
                epub_sequence=int(seq),
                chapter_num=chapter_num,
                chapter_full_title=title_raw,
                perk_name=name,
                classification=classification,
                jump=jump,
                cost=cost,
                cost_text=cost_text or "Free",
                free=free,
                perk_text=_norm(ws.cell(r, 7).value),
                constellation=_classify(constellation_idx, name, jump, overrides),
            )
        )
    return perks


# ---------- timeline --------------------------------------------------------


@dataclass
class TimelineEntry:
    sequence: int
    in_world_date_iso: str | None    # best-effort ISO date for sorting (None if range-only)
    in_world_date_text: str          # original cell value (preserved for display)
    events: str
    attribution: str                 # "author" | "Whamodyne"


_MONTH_NAMES = (
    "January|February|March|April|May|June|July|August|"
    "September|October|November|December"
)
# Pull out the first month name, optional first-day-of-range, and the
# year (which can appear several tokens later, e.g. "January-March, 2009"
# or "April 2-7, 2011" or "June, 2009-August 2010").
_FIRST_MONTH_RE = re.compile(rf"\b(?P<month>{_MONTH_NAMES})\b", re.IGNORECASE)
_FIRST_DAY_AFTER_MONTH_RE = re.compile(
    rf"\b(?:{_MONTH_NAMES})\s+(?P<day>\d{{1,2}})\b", re.IGNORECASE
)
_FIRST_YEAR_RE = re.compile(r"\b(?P<year>\d{4})\b")
_MONTHS = {
    m.lower(): i
    for i, m in enumerate(
        ("January February March April May June "
         "July August September October November December").split(),
        start=1,
    )
}


def _coerce_date(value) -> tuple[str | None, str]:
    """Return (sortable ISO if possible, original text)."""
    if isinstance(value, dt.datetime):
        iso = value.date().isoformat()
        return iso, iso
    if isinstance(value, dt.date):
        iso = value.isoformat()
        return iso, iso
    text = _norm(value)
    if not text:
        return None, ""
    # Best-effort: pull first month, first day-after-a-month, and first year
    # independently so range strings like "April 2-7, 2011" and
    # "June, 2009-August 2010" still produce a sortable ISO date.
    month_m = _FIRST_MONTH_RE.search(text)
    year_m = _FIRST_YEAR_RE.search(text)
    if not (month_m and year_m):
        return None, text
    day_m = _FIRST_DAY_AFTER_MONTH_RE.search(text)
    try:
        month = _MONTHS[month_m.group("month").lower()]
        day = int(day_m.group("day")) if day_m else 1
        iso = dt.date(int(year_m.group("year")), month, day).isoformat()
        return iso, text
    except (KeyError, ValueError):
        return None, text


def parse_timeline(wb) -> list[TimelineEntry]:
    ws = wb["Timeline of Events"]
    entries: list[TimelineEntry] = []
    seq = 0
    attribution = "author"   # column C resets at the Whamodyne marker
    for r in range(2, ws.max_row + 1):
        date_cell = ws.cell(r, 1).value
        events = _norm(ws.cell(r, 2).value)
        attr_cell = _norm(ws.cell(r, 3).value).lower()
        if "whamodyne" in attr_cell:
            attribution = "Whamodyne"
        elif "lord" in attr_cell:
            attribution = "author"
        if date_cell is None and not events:
            continue
        iso, text = _coerce_date(date_cell)
        seq += 1
        entries.append(
            TimelineEntry(
                sequence=seq,
                in_world_date_iso=iso,
                in_world_date_text=text,
                events=events,
                attribution=attribution,
            )
        )
    return entries


# ---------- main ------------------------------------------------------------


def main() -> None:
    wb = load_workbook(SRC, data_only=True)

    constellation_idx = _build_constellation_index(wb)
    overrides = _load_overrides()
    perks = parse_obtained_perks(wb, constellation_idx, overrides)

    classified = sum(1 for p in perks if p.constellation)
    write_validated_json(
        OUT_PERKS,
        {
            "_source": "data/raw/Brocktons_Celestial_Forge_Reference.xlsx#Obtained Perks",
            "_count": len(perks),
            "_coverage": "full story (EPUB sequences 1-192, story chapters 1 - 119.5)",
            "_classification_coverage": f"{classified}/{len(perks)} acquisitions classified to a constellation",
            "_note": (
                "Each row is one perk acquired. Cluster rolls and free-bonus perks "
                "each get their own row. cost is the numeric CP value (0 for any "
                "Free... variant); cost_text preserves the original wording. "
                "constellation is joined from perks_catalog.json + the Reference "
                "xlsx Unabridged List with cosmetic-tolerant matching; null when "
                "no match could be found."
            ),
            "perks": [asdict(p) for p in perks],
        },
        "obtained_perks",
    )

    entries = parse_timeline(wb)
    iso_dates = [e.in_world_date_iso for e in entries if e.in_world_date_iso]
    write_validated_json(
        OUT_TIMELINE,
        {
            "_source": "data/raw/Brocktons_Celestial_Forge_Reference.xlsx#Timeline of Events",
            "_attribution": (
                "Pre-story dates (through April 7, 2011) provided by the author "
                "(LordRoustabout). Story-period dates (April 8, 2011 onward) compiled "
                "by Whamodyne. Each entry is tagged in the `attribution` field."
            ),
            "_count": len(entries),
            "_first_in_world_date": min(iso_dates) if iso_dates else None,
            "_last_in_world_date": max(iso_dates) if iso_dates else None,
            "_note": (
                "in_world_date_iso is best-effort: ISO date for fully-specified rows, "
                "first-of-month for range rows like 'January-March, 2009', or null "
                "if no year. in_world_date_text preserves the original cell value."
            ),
            "entries": [asdict(e) for e in entries],
        },
        "timeline",
    )

    print(f"wrote {OUT_PERKS.relative_to(ROOT)}: {len(perks)} perks "
          f"across {len({p.chapter_num for p in perks})} chapters "
          f"({classified} classified, {len(perks) - classified} unclassified)")
    print(f"wrote {OUT_TIMELINE.relative_to(ROOT)}: {len(entries)} entries "
          f"({iso_dates and (min(iso_dates) + ' .. ' + max(iso_dates))})")


if __name__ == "__main__":
    main()
