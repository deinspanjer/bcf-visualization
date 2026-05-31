"""Build the canonical roll-attempt stream for downstream visualization.

This is the roll analogue of data/derived/timeline.json: it is the one
derived source `build_chapter_facts.py` should consume for roll attempts.

Inputs:
  - data/derived/rolls.json .............. curator roll log; wins where present
  - data/derived/roll_outcomes.json ...... interpolated fallback for uncovered chapters
  - data/derived/predicted_rolls.json .... mechanical predicted-slot schedule
  - data/manual/chapter_roll_overrides.json  manual per-chapter roll curation
  - data/manual/roll_overrides.json ...... optional row-level patches
  - data/derived/roll_text_evidence.json . predicted prose anchors
  - data/derived/perk_directory.json ..... canonical perk ids/constellations
  - data/derived/outstanding_perks_by_chapter.json  miss-size estimates

Output:
  - data/derived/roll_facts.json (validated)
"""

from __future__ import annotations

import datetime as _dt
import json
import re

from _common import write_validated_json
from data_paths import DERIVED, MANUAL, RAW, ROOT
from multi_grab import (
    load_overrides as load_multi_grab_overrides,
    merge_paid_units,
)
from perk_name_resolver import (
    build_alias_lookup,
    build_directory_match_index,
    load_perk_aliases,
)
from predict_rolls import _load_cp_words_per_chapter
from regime_simulator import (
    REGIMES,
    ShadowState,
    _accumulate_x100,
    load_regime_transitions,
    regime_for_chapter,
    regimes_for_chapter,
    shadow_words,
)
from roll_scheduler import (
    HitInput,
    SlotInput,
    diagnose_infeasible,
    schedule_chapter,
)

CURATOR_ROLLS = DERIVED / "rolls.json"
ROLL_OUTCOMES = DERIVED / "roll_outcomes.json"
PREDICTED_ROLLS = DERIVED / "predicted_rolls.json"
OBTAINED_PERKS = DERIVED / "obtained_perks.json"
CHAPTERS_JSON = DERIVED / "chapters.json"
EVIDENCE = DERIVED / "roll_text_evidence.json"
DIRECTORY = DERIVED / "perk_directory.json"
OUTSTANDING = DERIVED / "outstanding_perks_by_chapter.json"
ROLL_OVERRIDES = MANUAL / "roll_overrides.json"
SURVEY_DESIGNATIONS = RAW / "Brocktons_Celestial_Forge_Reference.xlsx"
OUT = DERIVED / "roll_facts.json"
VALIDATION_OUT = DERIVED / "roll_validation.json"


def sort_key_for_chapter(chapter_num: str) -> tuple[int, int]:
    parts = chapter_num.split(".")
    return (int(parts[0]), int(parts[1]) if len(parts) > 1 else 0)


def int_or_none(value) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def build_directory_lookup(directory: list[dict]):
    """Build the shared multi-key match index over the perk directory.

    Returned by callers as a single opaque ``DirectoryMatchIndex``;
    use ``lookup_perk(idx, ...)`` to query it.
    """
    aliases = load_perk_aliases()
    return build_directory_match_index(directory, aliases)


def lookup_perk(
    match_idx,
    name: str | None,
    jump: str | None = None,
    constellation: str | None = None,
) -> dict | None:
    if not name:
        return None
    return match_idx.lookup(name, jump=jump, constellation=constellation)


def _clean_designation_name(value: object) -> str:
    text = str(value or "").strip()
    return text.strip("\"'‘’“”").strip()


def _designation_key(name: str | None, constellation: str | None) -> tuple[str, str]:
    return (
        re.sub(r"\s+", " ", (name or "").strip()).casefold(),
        re.sub(r"\s+", " ", (constellation or "").strip()).casefold(),
    )


def _canonical_constellation_label(value: object) -> str:
    text = str(value or "").strip()
    if text.endswith(" Constellation"):
        text = text[: -len(" Constellation")]
    if text == "Crafting Skills":
        return "Crafting"
    return text


def _survey_designation_code(row: dict) -> str:
    parts = [
        str(row["order_obtained"]),
        str(row["constellation_code"]),
        str(row["constellation_order"]),
        str(row["cost_code"]),
    ]
    if row.get("bundle_order"):
        parts.append(str(row["bundle_order"]))
    return "-".join(parts)


def load_survey_designations(path=None) -> dict[tuple[str, str], dict]:
    """Load Survey's AI designation codes keyed by canonical perk identity."""
    path = path or SURVEY_DESIGNATIONS
    if not path or not path.exists():
        return {}

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    if "AI Perk Names" not in wb.sheetnames:
        return {}
    ws = wb["AI Perk Names"]

    constellation_names: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        label = row[9] if len(row) > 9 else None
        code = row[10] if len(row) > 10 else None
        if label and code:
            constellation_names[str(code).strip()] = _canonical_constellation_label(label)

    out: dict[tuple[str, str], dict] = {}
    source = "data/raw/Brocktons_Celestial_Forge_Reference.xlsx#AI Perk Names"
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        order, const_code, const_order, cost_code, bundle_order, name = row[:6]
        if order is None or not const_code or const_order is None or not cost_code or not name:
            continue
        designation = {
            "code": _survey_designation_code({
                "order_obtained": int(order),
                "constellation_code": str(const_code).strip(),
                "constellation_order": int(const_order),
                "cost_code": str(cost_code).strip(),
                "bundle_order": (
                    str(bundle_order).strip()
                    if bundle_order not in (None, "") else None
                ),
            }),
            "order_obtained": int(order),
            "constellation_code": str(const_code).strip(),
            "constellation_order": int(const_order),
            "cost_code": str(cost_code).strip(),
            "bundle_order": (
                str(bundle_order).strip()
                if bundle_order not in (None, "") else None
            ),
            "source": source,
        }
        clean_name = _clean_designation_name(name)
        constellation = constellation_names.get(str(const_code).strip(), "")
        out[_designation_key(clean_name, constellation)] = designation
        out.setdefault(_designation_key(clean_name, None), designation)
    return out


def survey_designation_for(
    designations: dict[tuple[str, str], dict],
    name: str | None,
    constellation: str | None,
) -> dict | None:
    if not name:
        return None
    return (
        designations.get(_designation_key(name, constellation))
        or designations.get(_designation_key(name, None))
    )


def normalized_evidence_kind(ev: dict | None, fallback: str) -> str:
    if ev is None:
        return fallback
    kind = ev.get("evidence_kind") or fallback
    return "silent" if kind == "no_evidence" else kind


def anchor_offset(ev: dict | None) -> int | None:
    if ev and ev.get("matching_events"):
        return ev["matching_events"][0]["anchor_offset"]
    return None


def next_cost_above(available_cp: int | None) -> int | None:
    if available_cp is None:
        return None
    for cost in (100, 200, 300, 400, 600, 800):
        if cost > available_cp:
            return cost
    return available_cp + 100


def miss_cost_estimate(
    outstanding_by_chapter: dict[str, dict[str, list[dict]]],
    chapter_num: str,
    constellation: str | None,
    available_cp: int | None,
) -> int | None:
    if available_cp is None:
        return None
    by_const = outstanding_by_chapter.get(chapter_num, {})
    candidates = by_const.get(constellation, []) if constellation else []
    if not candidates:
        candidates = [perk for perks in by_const.values() for perk in perks]
    costs = [
        int(perk["cost"])
        for perk in candidates
        if perk.get("cost") is not None and int(perk["cost"]) > available_cp
    ]
    return min(costs) if costs else next_cost_above(available_cp)


def perk_meta(
    match_idx,
    raw_perk: dict,
    parent_constellation: str | None,
) -> dict:
    """Resolve a raw curator perk dict to canonical directory meta.

    Returns ``{id, name, instance, jump, constellation, cost}``.

    - ``name`` is the canonical directory name (matches a wireframe
      star id). When the curator-typed raw string differs from the
      canonical, the raw string is preserved as ``instance`` for
      display; otherwise ``instance`` is None.
    - ``cost`` falls back to the raw curator cost only when no
      directory match was found.
    """
    raw_name = raw_perk.get("name") or raw_perk.get("perk_name")
    jump = raw_perk.get("source") or raw_perk.get("jump")
    constellation = raw_perk.get("constellation") or parent_constellation
    directory_meta = lookup_perk(match_idx, raw_name, jump, constellation)
    canonical_name = (
        directory_meta["name"] if directory_meta else (raw_name or "")
    )
    instance = raw_name if raw_name and raw_name != canonical_name else None
    return {
        "id": directory_meta["id"] if directory_meta else None,
        "name": canonical_name,
        "instance": instance,
        "jump": (
            directory_meta["jump"] if directory_meta
            else (jump or "unknown")
        ),
        "constellation": (
            (directory_meta.get("constellation") if directory_meta else None)
            or constellation
        ),
        "cost": (
            int_or_none(raw_perk.get("cost"))
            if raw_perk.get("cost") is not None
            else (directory_meta.get("cost") if directory_meta else None)
        ),
    }


def _purchased_perk_record(
    meta: dict,
    raw_perk: dict,
    survey_designations: dict[tuple[str, str], dict] | None = None,
) -> dict:
    """Build a canonical purchased_perks[*] row from a perk_meta result.

    Shape: ``{name, instance, id, cost, cost_unit, jump, free}``.
    ``name`` is the canonical directory name (matches a wireframe star);
    ``instance`` is the raw curator string when it differs (else null);
    ``id`` and ``jump`` come from the directory entry. ``cost_unit`` is
    propagated from the curator's parsed perk row (None for plain CP/WP
    costs; e.g. "Customization Points" for Zoids equipment where the
    pipeline multiplied the raw value by 100 to derive effective CP).
    """
    raw_name = raw_perk.get("name") or raw_perk.get("perk_name") or ""
    canonical = meta["name"] or raw_name
    instance = raw_name if raw_name and raw_name != canonical else None
    cost = (
        int(meta["cost"])
        if meta["cost"] is not None
        else int(raw_perk.get("cost") or 0)
    )
    record = {
        "name": canonical,
        "instance": instance,
        "id": meta["id"],
        "cost": cost,
        "cost_unit": raw_perk.get("cost_unit"),
        "jump": meta["jump"],
        "free": False,
    }
    designation = survey_designation_for(
        survey_designations or {},
        canonical,
        meta.get("constellation"),
    )
    if designation:
        record["survey_designation"] = designation
    return record


def split_hit_perks(
    match_idx,
    perks: list[dict],
    parent_constellation: str | None,
    survey_designations: dict[tuple[str, str], dict] | None = None,
) -> tuple[dict | None, list[dict], list[dict]]:
    """Returns (principal_meta, purchased_perks_array, free_perks_array).

    ``purchased_perks_array`` is shaped per ``_purchased_perk_record``.
    The principal is the first paid perk (preserves source order).
    """
    if not perks:
        return None, [], []
    paid_perks = [p for p in perks if not p.get("free", False)]
    free_perks_raw = [p for p in perks if p.get("free", False)]
    if not paid_perks:
        # All perks are free: still return one principal for scalar fields.
        principal = perks[0]
        principal_meta = perk_meta(match_idx, principal, parent_constellation)
        free_perks: list[dict] = []
        for perk in perks[1:]:
            free_meta = perk_meta(
                match_idx, perk,
                parent_constellation or principal_meta["constellation"],
            )
            free_perks.append({
                "id": free_meta["id"],
                "name": free_meta["name"],
                "jump": free_meta["jump"],
                "constellation": free_meta["constellation"],
            })
            designation = survey_designation_for(
                survey_designations or {},
                free_meta["name"],
                free_meta["constellation"],
            )
            if designation:
                free_perks[-1]["survey_designation"] = designation
        return principal_meta, [], free_perks
    principal = paid_perks[0]
    principal_meta = perk_meta(match_idx, principal, parent_constellation)
    purchased_perks: list[dict] = []
    for p in paid_perks:
        meta = perk_meta(match_idx, p, parent_constellation)
        purchased_perks.append(_purchased_perk_record(meta, p, survey_designations))
    free_perks: list[dict] = []
    for perk in free_perks_raw:
        free_meta = perk_meta(
            match_idx, perk,
            parent_constellation or principal_meta["constellation"],
        )
        free_perks.append({
            "id": free_meta["id"],
            "name": free_meta["name"],
            "jump": free_meta["jump"],
            "constellation": free_meta["constellation"],
        })
        designation = survey_designation_for(
            survey_designations or {},
            free_meta["name"],
            free_meta["constellation"],
        )
        if designation:
            free_perks[-1]["survey_designation"] = designation
    return principal_meta, purchased_perks, free_perks


def roll_base(
    *,
    roll_key: str,
    roll_number: int | None,
    chapter_num: str,
    predicted_chapter_num: str | None,
    source: str,
    source_kind: str,
    outcome: str,
    source_row_index: int,
    ev: dict | None,
) -> dict:
    return {
        "roll_key": roll_key,
        "roll_number": roll_number,
        "chapter_num": chapter_num,
        "predicted_chapter_num": predicted_chapter_num,
        "chapter_attribution_disagreement": (
            predicted_chapter_num is not None and predicted_chapter_num != chapter_num
        ),
        "source": source,
        "source_kind": source_kind,
        "outcome": outcome,
        "predicted_word_position_epub": (
            ev.get("cp_offset") if ev else None
        ),
        "epub_word_offset_predicted": (
            ev.get("epub_offset") if ev else None
        ),
        "epub_word_offset_curated": (
            ev.get("epub_offset") if ev else None
        ),
        "predicted_char_offset_in_chapter": (
            ev.get("predicted_char_offset") if ev else None
        ),
        "anchor_char_offset_in_chapter": anchor_offset(ev),
        "evidence_kind": normalized_evidence_kind(
            ev,
            "curator_log" if source == "curator_rolls" else "interpolated",
        ),
        "source_row_index": source_row_index,
    }


def _build_scheduler_inputs() -> dict[str, dict]:
    """Build per-chapter scheduler inputs: slots, hits, banked_cp_in,
    shadow_state, chapter_words, chapter_word_start.

    Walks chapters in canonical order. For each chapter we compute the
    starting banked CP and shadow state by replaying earlier chapters'
    "no-hits" simulation (hits are scheduler-decided per chapter, but for
    the inter-chapter ledger we use the SAME inputs predict_rolls used:
    proportional-spread hits, since obtained_perks gives us the hit
    counts and costs; even spread inside a chapter doesn't change the
    chapter-end banked CP if all hits land in-chapter and the scheduler
    is consistent).

    We use obtained_perks (paid only) as the canonical hit list per
    chapter; their order within a chapter is `epub_sequence` ascending.
    """
    chapters_doc = json.loads(CHAPTERS_JSON.read_text())
    perks_doc = json.loads(OBTAINED_PERKS.read_text())
    pred_doc = json.loads(PREDICTED_ROLLS.read_text())
    transitions = load_regime_transitions()
    multi_overrides = load_multi_grab_overrides()

    chapters = sorted(chapters_doc["chapters"], key=lambda c: tuple(c["sort_key"]))
    cp_words_by_title = _load_cp_words_per_chapter()
    chapter_words: dict[str, int] = {
        c["chapter_num"]: int(cp_words_by_title.get(c["full_title"], 0))
        for c in chapters
    }

    # Multi-grab merging: paid perks sharing (chapter, jump, epub_seq)
    # belong to the same Forge roll. Each merged unit produces ONE hit.
    perks_sorted = sorted(perks_doc["perks"], key=lambda p: p.get("epub_sequence", 0))
    units, _stats = merge_paid_units(perks_sorted, multi_overrides)
    units_by_chapter: dict[str, list[dict]] = {}
    for unit in units:
        units_by_chapter.setdefault(unit["chapter_num"], []).append(unit)

    # Group predicted rolls by chapter; word_position is story-global.
    pred_by_chapter: dict[str, list[dict]] = {}
    for r in pred_doc["predicted"]:
        pred_by_chapter.setdefault(r["chapter_num"], []).append(r)

    # Compute per-chapter starting word offsets (cumulative CP-earning words).
    chapter_word_start: dict[str, int] = {}
    cum = 0
    for c in chapters:
        chapter_word_start[c["chapter_num"]] = cum
        cum += chapter_words[c["chapter_num"]]

    # Walk to compute banked_cp_in / shadow_state at each chapter start.
    banked_in_per_chapter: dict[str, int] = {}
    shadow_in_per_chapter: dict[str, ShadowState] = {}
    segments_per_chapter: dict[str, list] = {}
    banked_x100 = 0
    shadow = ShadowState()
    from regime_simulator import shadow_words as _shadow_words
    for c in chapters:
        cn = c["chapter_num"]
        banked_in_per_chapter[cn] = banked_x100 // 100
        shadow_in_per_chapter[cn] = shadow.copy()
        # Walk this chapter's words at its regime(s), applying hits for
        # shadow bookkeeping. Hit positions: even-spread of merged units.
        words = chapter_words[cn]
        ch_units = units_by_chapter.get(cn, [])
        # Resolve regime segments for this chapter (mid-chapter transitions).
        segs = regimes_for_chapter(
            cn, transitions,
            [u["paid"][0] for u in ch_units] if ch_units else [],
            words,
        )
        segments_per_chapter[cn] = segs
        # Hit events: one per merged unit, position by even spread.
        if ch_units and words > 0:
            slot_w = words / len(ch_units)
            hit_events = sorted([
                (
                    int((i + 0.5) * slot_w),
                    sum(int(p.get("cost") or 0) for p in u["paid"]),
                    max(int(p.get("cost") or 0) for p in u["paid"]),
                )
                for i, u in enumerate(ch_units)
            ])
        else:
            hit_events = []
        last = 0
        seg_idx = 0
        current_regime = segs[seg_idx].regime

        def _walk(target: int) -> None:
            nonlocal banked_x100, last, seg_idx, current_regime
            while last < target:
                seg_end = segs[seg_idx].end_word_local
                if seg_end is None or seg_end >= target:
                    step = target - last
                    if step > 0:
                        banked_x100 = _accumulate_x100(
                            step, current_regime, banked_x100, shadow,
                        )
                    last = target
                    return
                step = seg_end - last
                if step > 0:
                    banked_x100 = _accumulate_x100(
                        step, current_regime, banked_x100, shadow,
                    )
                last = seg_end
                seg_idx += 1
                current_regime = segs[seg_idx].regime

        for hw, total_cost, principal_cost in hit_events:
            _walk(hw)
            banked_x100 -= total_cost * 100
            if banked_x100 < 0:
                banked_x100 = 0
            sw = _shadow_words(principal_cost, current_regime)
            if sw:
                shadow.remaining += sw
        if words > last:
            _walk(words)

    # Build scheduler inputs per chapter.
    out: dict[str, dict] = {}
    for c in chapters:
        cn = c["chapter_num"]
        words = chapter_words[cn]
        ws_global = chapter_word_start[cn]
        preds = pred_by_chapter.get(cn, [])
        preds.sort(key=lambda r: int(r["cp_offset"]))
        # Predicted slot list (chapter-local CP-word offset). This is
        # the model truth used for validation. The generation scheduler
        # may synthesize fallback slots below so roll_facts can still
        # carry source-derived rows, but validation keeps these counts
        # separate.
        predicted_slots = [
            SlotInput(
                word_position=max(0, int(r["cp_offset"]) - ws_global),
                roll_trigger_cp_threshold=int(
                    r.get("roll_trigger_cp_threshold")
                    or REGIMES[regime_for_chapter(cn)]["cp_per_roll"]
                ),
                source="predicted",
            )
            for r in preds
        ]
        slots = list(predicted_slots)
        # Hits = merged paid units in narrative order.
        ch_units = units_by_chapter.get(cn, [])
        hits = [
            HitInput(
                cost=sum(int(p.get("cost") or 0) for p in u["paid"]),
                perk=u["paid"][0],
                paid_perks=list(u["paid"]),
                free_perks=list(u["free_perks"]),
                mention_chapter_num=str(u.get("mention_chapter_num") or cn),
                mention_word_position=(
                    int(u["mention_word_position"])
                    if u.get("mention_word_position") is not None else None
                ),
                display_position_policy=(
                    u.get("display_position_policy") or "mechanical"
                ),
                evidence_quotes=list(u.get("evidence_quotes") or []),
            )
            for u in ch_units
        ]
        # If K > N, synthesize extra slots with even spread (matching
        # derive_roll_outcomes.py).
        n, k = len(slots), len(hits)
        if k > n and words > 0:
            extra = k - n
            # Evenly spaced at (i+1)/(extra+1) of words; collision-avoid
            # against existing slot positions.
            existing = {s.word_position for s in slots}
            for i in range(extra):
                pos = max(1, int(words * (i + 1) / (extra + 1)))
                while pos in existing:
                    pos += 1
                existing.add(pos)
                slots.append(SlotInput(
                    word_position=pos,
                    roll_trigger_cp_threshold=REGIMES[
                        regime_for_chapter(cn)
                    ]["cp_per_roll"],
                    source="synthetic",
                ))
            slots.sort(key=lambda s: s.word_position)
        out[cn] = {
            "slots": slots,
            "predicted_slots": predicted_slots,
            "synthetic_slot_count": len(slots) - len(predicted_slots),
            "hits": hits,
            "chapter_words": words,
            "chapter_word_start": ws_global,
            "banked_cp_in": banked_in_per_chapter.get(cn, 0),
            "shadow_in": shadow_in_per_chapter.get(cn, ShadowState()),
            "segments": segments_per_chapter.get(cn),
            "predicted_rolls": preds,
        }
    return out


def _norm_name(s: str | None) -> str:
    return (s or "").strip().lower()


def _is_metadata_only_roll_override(
    entry: dict, chapter_num: str | None = None
) -> bool:
    """True when an index-aligned override only carries non-structural metadata.

    Empty placeholder rows and quote-only rows should not rebuild the roll
    schedule. They patch the existing roll at the same chapter-local slot.
    """
    if not isinstance(entry, dict):
        return False
    if entry.get("outcome") in ("hit", "miss"):
        return False
    if entry.get("skipped") not in (None, False):
        return False
    if entry.get("source_ordinal") is not None:
        return False
    if entry.get("curator_added"):
        return False
    if entry.get("perks"):
        return False
    structural_fields = (
        "word_position",
    )
    if any(entry.get(field) not in (None, "", []) for field in structural_fields):
        return False
    return entry.get("display_position_policy") in (
        None,
        "",
        "mechanical",
        "mention",
        "section_start",
        "source_marker",
        "section_end",
    )


def _metadata_only_roll_override_has_manual_content(entry: dict) -> bool:
    return (
        bool(entry.get("constellation"))
        or bool(entry.get("evidence_quotes"))
        or entry.get("mention_word_position") is not None
        or entry.get("mention_chapter_num") is not None
    )


def _has_structural_roll_override(
    override: dict, chapter_num: str | None = None
) -> bool:
    return any(
        not _is_metadata_only_roll_override(entry, chapter_num)
        for entry in (override.get("rolls") or [])
    )


def _evidence_quotes(entry: dict | None) -> list[dict]:
    if not entry:
        return []
    quotes: list[dict] = []
    for quote in entry.get("evidence_quotes") or []:
        if not isinstance(quote, dict) or not quote.get("text"):
            continue
        payload = {
            "text": str(quote["text"]),
            "mention_chapter_num": (
                str(quote.get("mention_chapter_num"))
                if quote.get("mention_chapter_num") is not None else None
            ),
            "mention_word_position": (
                int(quote["mention_word_position"])
                if quote.get("mention_word_position") is not None else None
            ),
        }
        checkpoint = quote.get("cp_ledger_checkpoint")
        if isinstance(checkpoint, dict):
            banked_after = int_or_none(checkpoint.get("banked_cp_after_roll"))
            if (
                checkpoint.get("kind") == "post_roll_banked_cp_reset"
                and banked_after is not None
            ):
                payload["cp_ledger_checkpoint"] = {
                    "kind": "post_roll_banked_cp_reset",
                    "banked_cp_after_roll": banked_after,
                }
        quotes.append(payload)
    return quotes


def _first_quote_word_position(
    entry: dict | None, chapter_num: str,
) -> int | None:
    for quote in _evidence_quotes(entry):
        if str(quote.get("mention_chapter_num")) != str(chapter_num):
            continue
        word_position = quote.get("mention_word_position")
        if word_position is not None:
            return int(word_position)
    return None


def _apply_metadata(payload: dict, entry: dict | None, chapter_num: str) -> None:
    if not entry:
        return
    payload["_evidence_quotes"] = _evidence_quotes(entry)
    if entry.get("mention_chapter_num") is not None:
        payload["_mention_chapter_num"] = _norm_chapter(
            entry.get("mention_chapter_num"), chapter_num
        )
    if entry.get("mention_word_position") is not None:
        payload["_mention_word_position"] = entry.get("mention_word_position")
    if entry.get("display_position_policy") is not None:
        payload["_display_position_policy"] = entry.get("display_position_policy")
    if entry.get("curator_note") is not None:
        payload["_curator_note"] = entry.get("curator_note")
    if entry.get("constellation") is not None:
        payload["constellation"] = entry.get("constellation")
        payload["constellation_revealed"] = True


def _fallback_projection_override(entry: dict | None) -> dict | None:
    if not isinstance(entry, dict):
        return None
    if entry.get("skipped") not in (None, False):
        return None
    if entry.get("source_ordinal") is not None:
        return None
    if entry.get("curator_added"):
        return None
    if entry.get("perks"):
        return None
    if entry.get("word_position") not in (None, "", []):
        return None
    return entry


def _resolution_for_issue(override: dict | None, issue_code: str) -> dict | None:
    resolution = (override or {}).get("model_validation_resolution") or {}
    if resolution.get("status") != "resolved":
        return None
    resolved_codes = {
        str(code)
        for code in (resolution.get("resolved_issue_codes") or [])
    }
    if issue_code not in resolved_codes:
        return None
    return resolution


def _apply_issue_resolutions(
    issues: list[dict], override: dict | None
) -> tuple[list[dict], list[str]]:
    resolved_codes: list[str] = []
    resolved_issues: list[dict] = []
    for issue in issues:
        resolution = _resolution_for_issue(override, str(issue.get("code")))
        if resolution is None:
            resolved_issues.append(issue)
            continue
        resolved_codes.append(str(issue["code"]))
        patched = dict(issue)
        patched["severity"] = "info"
        patched["resolved"] = True
        patched["resolution_reason_code"] = resolution.get("reason_code")
        patched["resolution_note"] = resolution.get("note")
        resolved_issues.append(patched)
    return resolved_issues, resolved_codes


BLOCKING_VALIDATION_CODES = {
    "paid_rolls_exceed_predicted_slots",
    "known_attempts_exceed_predicted_slots",
    "cost_schedule_infeasible",
    "curated_hit_missing_perks",
}


def _validation_status(issues: list[dict]) -> tuple[str, bool, bool]:
    raw_has_discrepancy = any(
        issue["code"] in BLOCKING_VALIDATION_CODES for issue in issues
    )
    has_discrepancy = any(
        issue["code"] in BLOCKING_VALIDATION_CODES and not issue.get("resolved")
        for issue in issues
    )
    return (
        "discrepancy" if has_discrepancy else "ok",
        has_discrepancy,
        raw_has_discrepancy,
    )


def _manual_override_issues(chapter_num: str, override: dict | None) -> list[dict]:
    issues: list[dict] = []
    if not override:
        return issues
    for idx, entry in enumerate(override.get("rolls") or [], start=1):
        if not isinstance(entry, dict):
            continue
        if (
            entry.get("outcome") != "hit"
            or entry.get("perks")
            or entry.get("source_ordinal") is not None
        ):
            continue
        issues.append({
            "code": "curated_hit_missing_perks",
            "severity": "error",
            "message": f"Curated hit roll #{idx} does not name any perk.",
        })
    return issues


def _explicit_extra_slot_position(
    entry: dict | None, chapter_words: int
) -> int | None:
    if not isinstance(entry, dict):
        return None
    word_position = int_or_none(entry.get("word_position"))
    if word_position is not None:
        return word_position
    mention_word_position = int_or_none(entry.get("mention_word_position"))
    policy = entry.get("display_position_policy")
    if policy in {"mention", "source_marker"} and mention_word_position is not None:
        return mention_word_position
    if policy == "section_end":
        return max(0, int(chapter_words))
    return None


def _restructure_curator_rows(
    chapter_num: str,
    curator_rows: list[tuple[int, dict]],
    override: dict,
) -> list[dict]:
    """Restructure curator rows for an override-covered chapter.

    Returns a list of "synthetic curator-equivalent" row dicts in narrative
    order. Each row dict has the same shape as a curator row (kind,
    perks, banked_before, banked_after, constellation, raw, roll_number)
    plus extra keys ``_source_idx`` (carry-over) and
    ``_override_origin`` (set on rebuilt hit rows so downstream code can
    tell they came from the override).

    Algorithm:
      1. Walk curator rows; misses pass through unchanged.
      2. For each curator hit row, determine which override-rolls cover
         its perks (override roll perks ⊆ curator row's combined perks).
         Replace the curator hit row with the matching override-rolls,
         in override order. Walk banked_before/after across the
         override-rolls using the curator hit row's banked_before as the
         starting CP for the first override-roll, debiting paid costs.
      3. Sanity: the last override-roll's banked_after for the chapter
         should match curator's recorded banked_after for the chapter
         (last curator hit row that maps to a tail override-roll).
         Mismatch -> warn.
    """
    override_rolls_raw = override.get("rolls") or []
    metadata_only_by_index = {
        idx: entry
        for idx, entry in enumerate(override_rolls_raw)
        if _is_metadata_only_roll_override(entry, chapter_num)
    }
    structural_override_indices = [
        idx for idx, entry in enumerate(override_rolls_raw)
        if idx not in metadata_only_by_index
    ]

    def _passthrough_rows() -> list[dict]:
        out_rows: list[dict] = []
        non_trigger_index = 0
        for source_idx, row in curator_rows:
            payload = {
                "_source_idx": source_idx,
                "_override_origin": None,
                "kind": row.get("kind"),
                "perks": list(row.get("perks") or []),
                "banked_before": row.get("banked_before"),
                "banked_after": row.get("banked_after"),
                "constellation": row.get("constellation"),
                "constellation_revealed": row.get("constellation_revealed", False),
                "roll_number": row.get("roll_number"),
                "raw": row.get("raw"),
            }
            if row.get("kind") != "trigger":
                _apply_metadata(
                    payload, metadata_only_by_index.get(non_trigger_index), chapter_num
                )
                non_trigger_index += 1
            out_rows.append(payload)
        return out_rows

    if override_rolls_raw and not structural_override_indices:
        return _passthrough_rows()

    def _names_of(entry):
        return list(entry.get("perks") or [])
    override_rolls = {
        idx: _names_of(override_rolls_raw[idx])
        for idx in structural_override_indices
    }

    def _explicit_outcome(entry) -> str | None:
        if isinstance(entry, dict) and entry.get("outcome") in ("hit", "miss"):
            return entry.get("outcome")
        return None

    # Explicit miss-only overrides, such as Chapter 1's "both predicted
    # rolls are misses", should not be matched to a trigger or converted
    # into zero-cost hit rows. Preserve existing source rows unless the
    # override explicitly replaces them, and allow extra miss rows beyond
    # the source list so predicted-but-unlisted attempts can be curated.
    if structural_override_indices and all(
        _explicit_outcome(entry) == "miss" and not _names_of(entry)
        for entry in (override_rolls_raw[idx] for idx in structural_override_indices)
    ):
        out_rows: list[dict] = []
        non_trigger_templates = [
            (source_idx, row)
            for source_idx, row in curator_rows
            if row.get("kind") != "trigger"
        ]
        last_source_idx = non_trigger_templates[-1][0] if non_trigger_templates else 0
        for source_idx, row in curator_rows:
            if row.get("kind") != "trigger":
                continue
            out_rows.append({
                "_source_idx": source_idx,
                "_override_origin": None,
                "kind": "trigger",
                "perks": list(row.get("perks") or []),
                "banked_before": row.get("banked_before"),
                "banked_after": row.get("banked_after"),
                "constellation": row.get("constellation"),
                "constellation_revealed": row.get("constellation_revealed", False),
                "roll_number": row.get("roll_number"),
                "raw": row.get("raw"),
            })
        row_count = max(len(non_trigger_templates), len(override_rolls_raw))
        for idx in range(row_count):
            entry = override_rolls_raw[idx] if idx < len(override_rolls_raw) else None
            template = (
                non_trigger_templates[idx][1]
                if idx < len(non_trigger_templates) else {}
            )
            source_idx = (
                non_trigger_templates[idx][0]
                if idx < len(non_trigger_templates) else last_source_idx
            )
            if entry is None or _is_metadata_only_roll_override(entry, chapter_num):
                if not template:
                    continue
                payload = {
                    "_source_idx": source_idx,
                    "_override_origin": None,
                    "kind": template.get("kind"),
                    "perks": list(template.get("perks") or []),
                    "banked_before": template.get("banked_before"),
                    "banked_after": template.get("banked_after"),
                    "constellation": template.get("constellation"),
                    "constellation_revealed": template.get(
                        "constellation_revealed", False
                    ),
                    "roll_number": template.get("roll_number"),
                    "raw": template.get("raw"),
                }
                _apply_metadata(payload, entry, chapter_num)
                out_rows.append(payload)
                continue
            out_rows.append({
                "_source_idx": source_idx,
                "_override_origin": idx,
                "kind": "miss",
                "perks": [],
                "banked_before": template.get("banked_before"),
                "banked_after": template.get("banked_after"),
                "constellation": None,
                "constellation_revealed": False,
                "roll_number": (
                    template.get("roll_number")
                    if idx < len(non_trigger_templates) else None
                ),
                "raw": (
                    template.get("raw")
                    if idx < len(non_trigger_templates) else None
                ),
                "_evidence_quotes": _evidence_quotes(entry),
            })
        return out_rows
    # Normalize override rolls to (paid_names_set, all_names_in_order_with_meta)
    # We need cost info, but at this point we don't have it -- the caller
    # passes in curator perks (which carry cost). For each override roll,
    # we'll match by name against the host curator hit row's perks.

    # Build name -> perk dict lookup from ALL curator hit rows in this
    # chapter (so override names always resolve).
    name_to_perk: dict[str, dict] = {}
    for _, row in curator_rows:
        for p in row.get("perks") or []:
            nm = _norm_name(p.get("name"))
            if nm and nm not in name_to_perk:
                name_to_perk[nm] = p

    # Validate every override-roll perk is present in this chapter.
    for roll_idx, name_list in override_rolls.items():
        for nm in name_list:
            if _norm_name(nm) not in name_to_perk:
                raise ValueError(
                    f"multi_grab override for ch {chapter_num} roll #{roll_idx} "
                    f"references {nm!r} but no curator hit row in the chapter "
                    f"contains a perk by that name "
                    f"(known: {sorted(name_to_perk)})"
                )

    # For each override-roll, find which curator hit row "hosts" it
    # (i.e. the curator row whose perks superset all of the override-roll's
    # perks). If no single host covers it, fall back to the FIRST host
    # that covers any of its perks.
    def _curator_hit_index_for_override(
        ov_names: list[str], curator_rows: list[tuple[int, dict]]
    ) -> int:
        ov_set = {_norm_name(n) for n in ov_names}
        # First pass: full superset.
        for i, (_, row) in enumerate(curator_rows):
            if row.get("kind") in ("miss", "trigger"):
                continue
            row_set = {_norm_name((p.get("name") or "")) for p in row.get("perks") or []}
            if ov_set.issubset(row_set):
                return i
        # Fallback: any overlap.
        for i, (_, row) in enumerate(curator_rows):
            if row.get("kind") in ("miss", "trigger"):
                continue
            row_set = {_norm_name((p.get("name") or "")) for p in row.get("perks") or []}
            if ov_set & row_set:
                return i
        raise ValueError(
            f"multi_grab override for ch {chapter_num}: roll {ov_names!r} "
            f"could not be matched to any curator hit row"
        )

    # Map override-roll index -> host curator row index.
    override_to_host: dict[int, int] = {
        ov_idx: _curator_hit_index_for_override(name_list, curator_rows)
        for ov_idx, name_list in override_rolls.items()
    }

    # For each curator row, build the list of override-roll indices it hosts.
    host_to_overrides: dict[int, list[int]] = {}
    for ov_idx, host_idx in override_to_host.items():
        host_to_overrides.setdefault(host_idx, []).append(ov_idx)

    # Sanity: every curator hit row should be hosted by at least one
    # override-roll. (If not, the override silently drops perks; warn.)
    for i, (_, row) in enumerate(curator_rows):
        if row.get("kind") in ("miss", "trigger"):
            continue
        if i not in host_to_overrides:
            print(
                f"  derive_roll_facts: ch {chapter_num} override does not "
                f"cover curator hit row #{i} (perks: "
                f"{[(p.get('name')) for p in row.get('perks') or []]})"
            )

    # Walk curator rows; emit misses unchanged, replace each hit with the
    # override-rolls it hosts. Walk banked across the override-rolls using
    # the curator hit row's banked_before as the entry CP.
    out_rows: list[dict] = []
    non_trigger_index = 0
    for i, (source_idx, row) in enumerate(curator_rows):
        if row.get("kind") in ("miss", "trigger"):
            # Pass through (preserve curator banked_before/after).
            payload = {
                "_source_idx": source_idx,
                "_override_origin": None,
                "kind": row.get("kind"),
                "perks": list(row.get("perks") or []),
                "banked_before": row.get("banked_before"),
                "banked_after": row.get("banked_after"),
                "constellation": row.get("constellation"),
                "constellation_revealed": row.get("constellation_revealed", False),
                "roll_number": row.get("roll_number"),
                "raw": row.get("raw"),
            }
            if row.get("kind") != "trigger":
                _apply_metadata(
                    payload, metadata_only_by_index.get(non_trigger_index), chapter_num
                )
                non_trigger_index += 1
            out_rows.append(payload)
            continue

        ov_indices = host_to_overrides.get(i, [])
        if not ov_indices:
            # Override doesn't cover this curator hit row -- emit it
            # untouched as a fallback.
            payload = {
                "_source_idx": source_idx,
                "_override_origin": None,
                "kind": row.get("kind"),
                "perks": list(row.get("perks") or []),
                "banked_before": row.get("banked_before"),
                "banked_after": row.get("banked_after"),
                "constellation": row.get("constellation"),
                "constellation_revealed": row.get("constellation_revealed", False),
                "roll_number": row.get("roll_number"),
                "raw": row.get("raw"),
            }
            _apply_metadata(
                payload, metadata_only_by_index.get(non_trigger_index), chapter_num
            )
            non_trigger_index += 1
            out_rows.append(payload)
            continue

        host_before = int_or_none(row.get("banked_before"))
        host_after = int_or_none(row.get("banked_after"))
        running = host_before if host_before is not None else 0
        for j, ov_idx in enumerate(ov_indices):
            name_list = override_rolls[ov_idx]
            ov_perks: list[dict] = []
            ov_cost_total = 0
            for nm in name_list:
                src = name_to_perk[_norm_name(nm)]
                ov_perks.append({
                    "name": src.get("name"),
                    "source": src.get("source"),
                    "cost": src.get("cost", 0),
                    "free": bool(src.get("free", False)),
                    "constellation": src.get("constellation"),
                })
                if not src.get("free", False):
                    ov_cost_total += int(src.get("cost") or 0)
            before = running
            after = before - ov_cost_total
            running = after
            # Last override-roll for this host inherits curator's roll_number /
            # raw / constellation_revealed; earlier ones get None roll_number
            # so downstream evidence lookup doesn't double-attach.
            is_last_for_host = (j == len(ov_indices) - 1)
            override_entry = override_rolls_raw[ov_idx]
            override_constellation = (
                override_entry.get("constellation")
                if isinstance(override_entry, dict) else None
            )
            out_rows.append({
                "_source_idx": source_idx,
                "_override_origin": ov_idx,
                "kind": "roll",
                "perks": ov_perks,
                "banked_before": before,
                "banked_after": after,
                "constellation": (
                    override_constellation
                    or (ov_perks[0].get("constellation") if ov_perks else None)
                    or row.get("constellation")
                ),
                "constellation_revealed": (
                    row.get("constellation_revealed", False)
                    if is_last_for_host else False
                ),
                "roll_number": (
                    row.get("roll_number") if is_last_for_host else None
                ),
                "raw": row.get("raw") if is_last_for_host else None,
                "_evidence_quotes": _evidence_quotes(override_rolls_raw[ov_idx]),
            })
        non_trigger_index += 1
        # Sanity: final running balance should equal host_after for the
        # last override hosted by this curator row.
        if host_after is not None and running != host_after:
            print(
                f"  derive_roll_facts: ch {chapter_num} override walk "
                f"banked_after mismatch for curator hit row #{i}: "
                f"override={running} curator={host_after} (rolls "
                f"{[override_rolls[k] for k in ov_indices]})"
            )

    return out_rows


def _norm_chapter(value: str | None, fallback: str) -> str:
    return str(value) if value is not None else str(fallback)


def _override_needs_direct_rows(
    chapter_num: str,
    override: dict,
    curator_rows: list[tuple[int, dict]] | None = None,
) -> bool:
    """True when the override cannot be matched only against same-chapter
    curator hit rows.

    Existing multi-grab overrides split curator rows within a chapter.
    Cross-chapter evidence can point at a later mention/listing chapter,
    so those overrides need direct construction from obtained_perks instead.
    """
    for entry in override.get("rolls") or []:
        if isinstance(entry, dict) and entry.get("skipped"):
            return True
        if isinstance(entry, dict) and entry.get("source_ordinal") is not None:
            return True
        if _norm_chapter(entry.get("mention_chapter_num"), chapter_num) != str(chapter_num):
            return True
    if curator_rows is not None:
        structural_entries = [
            entry for entry in override.get("rolls") or []
            if not _is_metadata_only_roll_override(entry, chapter_num)
        ]
        if (
            any(entry.get("outcome") == "miss" for entry in structural_entries)
            and any(entry.get("outcome") == "hit" for entry in structural_entries)
        ):
            return True
        if any(
            entry.get("outcome") == "hit" and not entry.get("perks")
            for entry in structural_entries
        ):
            return True
        hosted_names = {
            _norm_name(p.get("name"))
            for _idx, row in curator_rows
            if row.get("kind") not in ("miss", "trigger")
            for p in row.get("perks") or []
        }
        for entry in override.get("rolls") or []:
            if _is_metadata_only_roll_override(entry, chapter_num):
                continue
            for name in entry.get("perks") or []:
                if _norm_name(name) not in hosted_names:
                    return True
    return False


def _build_obtained_lookup(obtained_perks: list[dict]) -> dict[tuple[str, str], list[dict]]:
    lookup: dict[tuple[str, str], list[dict]] = {}
    for perk in obtained_perks:
        name = perk.get("perk_name") or perk.get("name")
        if not name:
            continue
        lookup.setdefault((str(perk["chapter_num"]), _norm_name(name)), []).append(perk)
    return lookup


def _source_roll_occurrences(
    curator_rolls: list[dict],
    multi_overrides: dict,
    chapter_word_start_global: dict[str, int] | None = None,
) -> tuple[dict[int, dict], dict[tuple[str, int], dict], dict[int, dict]]:
    chapter_word_start_global = chapter_word_start_global or {}
    non_trigger_seq_by_chapter: dict[str, int] = {}
    by_raw_roll_number: dict[int, dict] = {}
    by_chapter_index: dict[tuple[str, int], dict] = {}
    by_source_ordinal: dict[int, dict] = {}
    source_ordinal = 0
    for source_row_index, row in enumerate(curator_rolls):
        if row.get("kind") not in {"roll", "miss"}:
            continue
        roll_number = row.get("roll_number")
        if roll_number is None:
            continue
        source_ordinal += 1
        chapter_num = str(row.get("chapter_num"))
        seq = non_trigger_seq_by_chapter.get(chapter_num, 0) + 1
        non_trigger_seq_by_chapter[chapter_num] = seq
        override_rolls = (
            (multi_overrides.get(chapter_num) or {}).get("rolls") or []
        )
        override = (
            override_rolls[seq - 1]
            if 0 <= seq - 1 < len(override_rolls)
            and isinstance(override_rolls[seq - 1], dict)
            else None
        )
        word_position = None
        if override is not None and override.get("mention_word_position") is not None:
            word_position = int(override["mention_word_position"])
        if word_position is None:
            word_position = _first_quote_word_position(override, chapter_num)
        cumulative = (
            chapter_word_start_global.get(chapter_num, 0) + word_position
            if word_position is not None else None
        )
        info = {
            "source_row_index": int(source_row_index),
            "source_ordinal": source_ordinal,
            "source_chapter_num": chapter_num,
            "source_chapter_ordinal": seq,
            "source_roll_label": f"Roll {int(roll_number)}",
            "source_word_position": word_position,
            "source_cumulative_word_offset": cumulative,
        }
        by_raw_roll_number.setdefault(int(roll_number), info)
        by_chapter_index[(chapter_num, seq)] = info
        by_source_ordinal[source_ordinal] = info
    return by_raw_roll_number, by_chapter_index, by_source_ordinal


def _visible_chapter_nums(*values: str | None) -> list[str]:
    chapters = {str(value) for value in values if value is not None}
    return sorted(chapters, key=sort_key_for_chapter)


def _obtained_as_roll_perk(perk: dict) -> dict:
    return {
        "name": perk.get("perk_name") or perk.get("name"),
        "source": perk.get("jump") or perk.get("source"),
        "cost": int(perk.get("cost") or 0),
        "free": bool(perk.get("free", False)),
        "constellation": perk.get("constellation"),
    }


def _direct_override_rows(
    chapter_num: str,
    curator_rows: list[tuple[int, dict]],
    override: dict,
    obtained_lookup: dict[tuple[str, str], list[dict]],
    source_templates_by_source_ordinal: dict[int, tuple[int, dict]] | None = None,
) -> list[dict]:
    """Build curator-equivalent rows directly from an index-aligned
    chapter override.

    This path is for deferred rows where the mechanical chapter and the
    mention/listing chapter differ, so no same-chapter curator row can
    host the perk.
    """
    out_rows: list[dict] = []
    non_trigger_templates = [
        (source_idx, row)
        for source_idx, row in curator_rows
        if row.get("kind") != "trigger"
    ]
    source_position_by_row_index = {
        source_idx: position
        for position, (source_idx, _row) in enumerate(non_trigger_templates)
    }
    for source_idx, row in curator_rows:
        if row.get("kind") != "trigger":
            continue
        out_rows.append({
            "_source_idx": source_idx,
            "_override_origin": None,
            "_override_direct": False,
            "kind": "trigger",
            "perks": list(row.get("perks") or []),
            "banked_before": row.get("banked_before"),
            "banked_after": row.get("banked_after"),
            "constellation": row.get("constellation"),
            "constellation_revealed": row.get("constellation_revealed", False),
            "roll_number": row.get("roll_number"),
            "raw": row.get("raw"),
        })

    last_source_idx = non_trigger_templates[-1][0] if non_trigger_templates else 0
    override_rolls = override.get("rolls") or []
    source_cursor = 0
    source_templates_by_source_ordinal = source_templates_by_source_ordinal or {}
    explicit_source_positions: set[int] = set()
    for entry in override_rolls:
        source_ordinal = (
            entry.get("source_ordinal")
            if isinstance(entry, dict) else None
        )
        if source_ordinal is None:
            continue
        found = source_templates_by_source_ordinal.get(int(source_ordinal))
        if found is None:
            continue
        position = source_position_by_row_index.get(int(found[0]))
        if position is not None:
            explicit_source_positions.add(position)
    consumed_source_positions: set[int] = set(explicit_source_positions)

    def _next_source_template() -> tuple[int | None, dict, bool]:
        nonlocal source_cursor
        skipped_explicit_source = False
        while (
            source_cursor < len(non_trigger_templates)
            and source_cursor in consumed_source_positions
        ):
            if source_cursor in explicit_source_positions:
                skipped_explicit_source = True
            source_cursor += 1
        if source_cursor >= len(non_trigger_templates):
            return last_source_idx, {}, skipped_explicit_source
        consumed_source_positions.add(source_cursor)
        source_idx, template = non_trigger_templates[source_cursor]
        source_cursor += 1
        return source_idx, template, skipped_explicit_source

    def _source_template_for_entry(
        entry: dict | None,
    ) -> tuple[int | None, dict, bool]:
        source_ordinal = (
            entry.get("source_ordinal")
            if isinstance(entry, dict) else None
        )
        if source_ordinal is not None:
            found = source_templates_by_source_ordinal.get(int(source_ordinal))
            if found is not None:
                position = source_position_by_row_index.get(int(found[0]))
                if position is not None:
                    consumed_source_positions.add(position)
                return found[0], found[1], False
        return _next_source_template()

    for override_idx, entry in enumerate(override_rolls):
        if isinstance(entry, dict) and entry.get("skipped"):
            out_rows.append({
                "_source_idx": last_source_idx,
                "_override_origin": override_idx,
                "_override_direct": True,
                "kind": "skipped",
                "perks": [],
                "banked_before": None,
                "banked_after": None,
                "constellation": None,
                "constellation_revealed": False,
                "roll_number": None,
                "raw": None,
            })
            continue
        if entry is None or _is_metadata_only_roll_override(entry, chapter_num):
            source_idx, template, skipped_explicit_source = (
                _source_template_for_entry(entry)
            )
            if not template:
                if (
                    skipped_explicit_source and isinstance(entry, dict)
                    and _metadata_only_roll_override_has_manual_content(entry)
                ):
                    display_policy = entry.get("display_position_policy")
                    if display_policy is None:
                        display_policy = "mechanical"
                    out_rows.append({
                        "_source_idx": source_idx,
                        "_override_origin": override_idx,
                        "_override_direct": True,
                        "_curator_added": False,
                        "_mention_chapter_num": _norm_chapter(
                            entry.get("mention_chapter_num"), chapter_num
                        ),
                        "_mention_word_position": entry.get("mention_word_position"),
                        "_display_position_policy": display_policy,
                        "_evidence_quotes": _evidence_quotes(entry),
                        "kind": "miss",
                        "perks": [],
                        "banked_before": None,
                        "banked_after": None,
                        "constellation": entry.get("constellation"),
                        "constellation_revealed": bool(entry.get("constellation")),
                        "roll_number": None,
                        "raw": None,
                        "_source_identity_inferred": False,
                    })
                elif isinstance(entry, dict):
                    out_rows.append({
                        "_source_idx": source_idx,
                        "_override_origin": override_idx,
                        "_override_direct": True,
                        "kind": "skipped",
                        "perks": [],
                        "banked_before": None,
                        "banked_after": None,
                        "constellation": None,
                        "constellation_revealed": False,
                        "roll_number": None,
                        "raw": None,
                    })
                continue
            payload = {
                "_source_idx": source_idx,
                "_override_origin": None,
                "_override_direct": False,
                "kind": template.get("kind"),
                "perks": list(template.get("perks") or []),
                "banked_before": template.get("banked_before"),
                "banked_after": template.get("banked_after"),
                "constellation": template.get("constellation"),
                "constellation_revealed": template.get("constellation_revealed", False),
                "roll_number": template.get("roll_number"),
                "raw": template.get("raw"),
                "_source_identity_inferred": True,
            }
            if isinstance(entry, dict):
                if entry.get("source_ordinal") is not None:
                    payload["_source_ordinal"] = int(entry["source_ordinal"])
            _apply_metadata(payload, entry, chapter_num)
            out_rows.append(payload)
            continue
        if (
            entry.get("curator_added")
            and entry.get("source_ordinal") is None
            and entry.get("outcome") == "miss"
            and not entry.get("perks")
        ):
            source_idx, template = last_source_idx, {}
            curator_added = True
        else:
            source_idx, template, _skipped_explicit_source = _source_template_for_entry(entry)
            curator_added = False
        has_source_template = bool(template)
        mention_chapter = _norm_chapter(entry.get("mention_chapter_num"), chapter_num)
        template_kind = template.get("kind")
        outcome = entry.get("outcome") or (
            "hit" if template_kind == "roll"
            else "miss" if template_kind == "miss"
            else "hit" if entry.get("perks") else "miss"
        )
        roll_perks: list[dict] = []
        matched_obtained_perks: list[dict] = []
        if entry.get("perks"):
            for name in entry.get("perks") or []:
                candidates = obtained_lookup.get((mention_chapter, _norm_name(name))) or []
                if not candidates:
                    candidates = obtained_lookup.get((chapter_num, _norm_name(name))) or []
                if not candidates:
                    raise ValueError(
                        f"chapter_roll_overrides ch {chapter_num} roll #{override_idx} "
                        f"references {name!r}, but obtained_perks has no matching "
                        f"perk in chapter {mention_chapter} or {chapter_num}"
                    )
                matched = candidates[0]
                matched_obtained_perks.append(matched)
                roll_perks.append(_obtained_as_roll_perk(matched))
            seen_free_names = {
                _norm_name(perk.get("name"))
                for perk in roll_perks
                if perk.get("free")
            }
            all_obtained = [
                perk
                for candidates in obtained_lookup.values()
                for perk in candidates
            ]
            for matched in matched_obtained_perks:
                for perk in all_obtained:
                    if not perk.get("free"):
                        continue
                    if str(perk.get("chapter_num")) != str(matched.get("chapter_num")):
                        continue
                    if perk.get("epub_sequence") != matched.get("epub_sequence"):
                        continue
                    free_name = _norm_name(perk.get("perk_name") or perk.get("name"))
                    if free_name in seen_free_names:
                        continue
                    seen_free_names.add(free_name)
                    roll_perks.append(_obtained_as_roll_perk(perk))
        elif outcome == "hit":
            roll_perks = list(template.get("perks") or [])

        display_policy = entry.get("display_position_policy")
        if display_policy is None:
            display_policy = "mechanical"
        explicit_source_ordinal = (
            int(entry["source_ordinal"])
            if entry.get("source_ordinal") is not None else None
        )
        payload = {
            "_source_idx": source_idx,
            "_override_origin": override_idx,
            "_override_direct": True,
            "_curator_added": curator_added,
            "_mention_chapter_num": mention_chapter,
            "_mention_word_position": entry.get("mention_word_position"),
            "_display_position_policy": display_policy,
            "_evidence_quotes": _evidence_quotes(entry),
            "kind": "miss" if outcome == "miss" else "roll",
            "perks": roll_perks,
            "banked_before": template.get("banked_before"),
            "banked_after": template.get("banked_after"),
            "constellation": (
                entry.get("constellation")
                or (roll_perks[0].get("constellation") if roll_perks else None)
                or template.get("constellation")
            ),
            "constellation_revealed": bool(entry.get("constellation")),
            "roll_number": template.get("roll_number"),
            "raw": template.get("raw"),
            "_source_identity_inferred": has_source_template,
        }
        if explicit_source_ordinal is not None:
            payload["_source_ordinal"] = explicit_source_ordinal
            payload["_source_identity_inferred"] = True
        out_rows.append(payload)
    for pos, (source_idx, template) in enumerate(non_trigger_templates):
        if pos in consumed_source_positions:
            continue
        payload = {
            "_source_idx": source_idx,
            "_override_origin": None,
            "_override_direct": False,
            "kind": template.get("kind"),
            "perks": list(template.get("perks") or []),
            "banked_before": template.get("banked_before"),
            "banked_after": template.get("banked_after"),
            "constellation": template.get("constellation"),
            "constellation_revealed": template.get("constellation_revealed", False),
            "roll_number": template.get("roll_number"),
            "raw": template.get("raw"),
            "_source_identity_inferred": True,
        }
        out_rows.append(payload)
    return out_rows


def _apply_overrides(rolls: list[dict], overrides_doc: dict) -> int:
    """Apply hand-curation overrides as a final pass. Returns count."""
    roll_overrides = overrides_doc.get("roll_overrides") or {}
    applied = 0
    for r in rolls:
        patch = roll_overrides.get(r.get("roll_key"))
        if not patch:
            continue
        for k, v in patch.items():
            if k.startswith("_"):
                continue
            r[k] = v
        r["slot_source"] = "override"
        applied += 1
    return applied


def _forge_cp_debit_total(roll: dict) -> int:
    total = 0
    for perk in roll.get("purchased_perks") or []:
        if perk.get("cost") is None:
            continue
        total += int(perk["cost"])
    return total


def _roll_accounting_locator(index: int, roll: dict) -> dict:
    return {
        "roll_ordinal": roll.get("roll_ordinal"),
        "roll_label": roll.get("roll_label"),
        "predicted_ordinal": roll.get("predicted_ordinal"),
        "predicted_label": roll.get("predicted_label"),
        "source_ordinal": roll.get("source_ordinal"),
        "source_label": roll.get("source_label"),
        "mechanical_chapter_num": roll.get("mechanical_chapter_num"),
        "row_index": index,
        "roll_key": roll.get("roll_key"),
    }


def _normalize_hit_accounting(rolls: list[dict]) -> tuple[list[dict], list[dict]]:
    corrections: list[dict] = []
    missing_available: list[dict] = []
    for index, roll in enumerate(rolls):
        if roll.get("outcome") != "hit":
            continue
        debit = _forge_cp_debit_total(roll)
        if debit <= 0:
            continue
        available_cp = int_or_none(roll.get("available_cp"))
        if available_cp is None:
            missing_available.append(_roll_accounting_locator(index, roll) | {
                "debit": debit,
                "available_cp": roll.get("available_cp"),
                "banked_cp_after_roll": roll.get("banked_cp_after_roll"),
            })
            continue
        expected = max(0, available_cp - debit)
        if roll.get("banked_cp_after_roll") != expected:
            old = roll.get("banked_cp_after_roll")
            roll["banked_cp_after_roll"] = expected
            corrections.append(_roll_accounting_locator(index, roll) | {
                "available_cp": available_cp,
                "debit": debit,
                "old_banked_cp_after_roll": old,
                "new_banked_cp_after_roll": expected,
            })
    return corrections, missing_available


def _principal_forge_cp_cost(roll: dict) -> int:
    costs = [
        int(perk["cost"])
        for perk in roll.get("purchased_perks") or []
        if perk.get("cost") is not None
    ]
    return max(costs) if costs else 0


def _ledger_checkpoint_for_roll(roll: dict) -> dict | None:
    for quote in roll.get("evidence_quotes") or []:
        if not isinstance(quote, dict):
            continue
        checkpoint = quote.get("cp_ledger_checkpoint")
        if not isinstance(checkpoint, dict):
            continue
        if checkpoint.get("kind") != "post_roll_banked_cp_reset":
            continue
        if int_or_none(checkpoint.get("banked_cp_after_roll")) is None:
            continue
        return dict(checkpoint)
    return None


def _accumulate_ledger_span(
    *,
    start_global: int,
    end_global: int,
    banked_cp_x100: int,
    shadow_state: ShadowState,
    chapter_order: list[str],
    scheduler_inputs: dict[str, dict],
) -> int:
    if end_global <= start_global:
        return banked_cp_x100
    cursor = int(start_global)
    for chapter_num in chapter_order:
        inp = scheduler_inputs.get(chapter_num)
        if inp is None:
            continue
        chapter_start = int(inp.get("chapter_word_start") or 0)
        chapter_words = int(inp.get("chapter_words") or 0)
        chapter_end = chapter_start + chapter_words
        if cursor >= end_global:
            break
        if end_global <= chapter_start or cursor >= chapter_end:
            continue
        local_start = max(0, cursor - chapter_start)
        local_end = min(end_global, chapter_end) - chapter_start
        if local_end <= local_start:
            continue
        segment_start = 0
        for segment in inp.get("segments") or []:
            segment_end = (
                int(segment.end_word_local)
                if segment.end_word_local is not None else chapter_words
            )
            span_start = max(local_start, segment_start)
            span_end = min(local_end, segment_end)
            if span_end > span_start:
                banked_cp_x100 = _accumulate_x100(
                    span_end - span_start,
                    int(segment.regime),
                    banked_cp_x100,
                    shadow_state,
                )
            segment_start = segment_end
            if segment_start >= local_end:
                break
        cursor = chapter_start + local_end
    return banked_cp_x100


def _restamp_roll_accounting(
    roll: dict,
    *,
    available_cp: int,
    outstanding_by_chapter: dict[str, dict],
) -> int:
    roll["available_cp"] = int(available_cp)
    if roll.get("outcome") == "hit":
        after = max(0, int(available_cp) - _forge_cp_debit_total(roll))
        roll["banked_cp_after_roll"] = after
        return after

    roll["banked_cp_after_roll"] = int(available_cp)
    estimate = miss_cost_estimate(
        outstanding_by_chapter,
        str(roll.get("mechanical_chapter_num") or roll.get("chapter_num")),
        roll.get("constellation"),
        int(available_cp),
    )
    roll["miss_cost_estimate"] = estimate
    roll["rolled_perk_cost"] = estimate
    return int(available_cp)


def _apply_cp_ledger_checkpoints(
    rolls: list[dict],
    *,
    chapter_order: list[str],
    scheduler_inputs: dict[str, dict],
    outstanding_by_chapter: dict[str, dict],
) -> list[dict]:
    applications: list[dict] = []
    ordered_indices = sorted(
        range(len(rolls)),
        key=lambda index: (
            rolls[index].get("mechanical_cumulative_word_offset")
            if rolls[index].get("mechanical_cumulative_word_offset") is not None
            else 10**18,
            index,
        ),
    )
    ledger_active = False
    cursor_global = 0
    banked_cp_x100 = 0
    shadow_state = ShadowState()

    for index in ordered_indices:
        roll = rolls[index]
        if roll.get("source_kind") == "trigger":
            continue
        coord = int_or_none(roll.get("mechanical_cumulative_word_offset"))
        if ledger_active and coord is not None:
            banked_cp_x100 = _accumulate_ledger_span(
                start_global=cursor_global,
                end_global=coord,
                banked_cp_x100=banked_cp_x100,
                shadow_state=shadow_state,
                chapter_order=chapter_order,
                scheduler_inputs=scheduler_inputs,
            )
            available = banked_cp_x100 // 100
            after = _restamp_roll_accounting(
                roll,
                available_cp=available,
                outstanding_by_chapter=outstanding_by_chapter,
            )
            banked_cp_x100 = after * 100
            if roll.get("outcome") == "hit":
                regime = regime_for_chapter(
                    str(roll.get("mechanical_chapter_num") or roll.get("chapter_num"))
                )
                shadow_state.remaining += shadow_words(
                    _principal_forge_cp_cost(roll),
                    regime,
                )
            cursor_global = coord

        checkpoint = _ledger_checkpoint_for_roll(roll)
        if checkpoint is None:
            continue
        reset_after = int(checkpoint["banked_cp_after_roll"])
        debit = _forge_cp_debit_total(roll)
        new_available = (
            reset_after + debit if roll.get("outcome") == "hit" else reset_after
        )
        application = _roll_accounting_locator(index, roll) | {
            "old_available_cp": roll.get("available_cp"),
            "old_banked_cp_after_roll": roll.get("banked_cp_after_roll"),
            "new_available_cp": new_available,
            "new_banked_cp_after_roll": reset_after,
            "checkpoint": checkpoint,
        }
        roll["available_cp"] = new_available
        roll["banked_cp_after_roll"] = reset_after
        if roll.get("outcome") == "miss":
            estimate = miss_cost_estimate(
                outstanding_by_chapter,
                str(roll.get("mechanical_chapter_num") or roll.get("chapter_num")),
                roll.get("constellation"),
                new_available,
            )
            roll["miss_cost_estimate"] = estimate
            roll["rolled_perk_cost"] = estimate
        applications.append(application)

        ledger_active = coord is not None
        if ledger_active:
            cursor_global = int(coord)
            banked_cp_x100 = reset_after * 100
            shadow_state = ShadowState()
            if roll.get("outcome") == "hit":
                regime = regime_for_chapter(
                    str(roll.get("mechanical_chapter_num") or roll.get("chapter_num"))
                )
                shadow_state.remaining += shadow_words(
                    _principal_forge_cp_cost(roll),
                    regime,
                )

    return applications


def _ordinal_label(prefix: str, value: int | None) -> str | None:
    return f"{prefix}{value}" if value is not None else None


def _record_roll_identity(
    record: dict,
    *,
    predicted_ordinal: int | None,
    source_ordinal: int | None,
    association_source: str,
) -> None:
    record["_predicted_ordinal"] = predicted_ordinal
    record["_source_ordinal"] = source_ordinal
    record["association_source"] = association_source


def _association_source_for(
    *,
    predicted_ordinal: int | None,
    source_ordinal: int | None,
    curated: bool,
) -> str:
    if predicted_ordinal is None or source_ordinal is None:
        return "none"
    return "curated" if curated else "auto"


def _owner_chapter_sort_key(chapter_order: list[str], chapter_num: str) -> tuple[int, tuple[int, int]]:
    try:
        return (chapter_order.index(chapter_num), (0, 0))
    except ValueError:
        return (10**9, sort_key_for_chapter(chapter_num))


def _stamp_explicit_ordinals(rolls: list[dict], chapter_order: list[str]) -> None:
    def roll_owner_chapter(roll: dict) -> str:
        return str(roll.get("mechanical_chapter_num") or roll["chapter_num"])

    def roll_owner_position(roll: dict) -> int:
        value = (
            roll.get("mechanical_cumulative_word_offset")
            if roll.get("mechanical_cumulative_word_offset") is not None
            else roll.get("display_cumulative_word_offset")
        )
        return int(value) if value is not None else 10**12

    rolls_by_owner: dict[str, list[dict]] = {}
    for roll in rolls:
        rolls_by_owner.setdefault(roll_owner_chapter(roll), []).append(roll)

    ordered_rolls: list[dict] = []
    ordered_all_rolls: list[dict] = []
    for owner_chapter, owner_rolls in sorted(
        rolls_by_owner.items(),
        key=lambda item: _owner_chapter_sort_key(chapter_order, item[0]),
    ):
        owner_rolls.sort(key=lambda r: (
            0 if r.get("source_kind") == "trigger" else 1,
            roll_owner_position(r),
            r.get("source_ordinal") if r.get("source_ordinal") is not None else 10**12,
            r.get("source_row_index", 0),
            r.get("roll_key", ""),
        ))
        total_owner_rolls = len(owner_rolls)
        chapter_ordinal = 0
        for seq, roll in enumerate(owner_rolls, start=1):
            roll["roll_sequence_in_chapter"] = seq
            roll["rolls_in_chapter"] = total_owner_rolls
            if roll.get("source_kind") == "trigger":
                roll["chapter_ordinal"] = None
                roll["chapter_label"] = None
                ordered_all_rolls.append(roll)
                continue
            chapter_ordinal += 1
            roll["chapter_ordinal"] = chapter_ordinal
            roll["chapter_label"] = _ordinal_label("C", chapter_ordinal)
            ordered_rolls.append(roll)
            ordered_all_rolls.append(roll)

    for roll_ordinal, roll in enumerate(ordered_rolls, start=1):
        roll["roll_ordinal"] = roll_ordinal
        roll["roll_label"] = _ordinal_label("R", roll_ordinal)

    for roll in rolls:
        predicted_ordinal = int_or_none(roll.pop("_predicted_ordinal", None))
        source_ordinal = int_or_none(roll.pop("_source_ordinal", None))
        roll.setdefault("roll_ordinal", None)
        roll.setdefault("roll_label", None)
        roll.setdefault("chapter_ordinal", None)
        roll.setdefault("chapter_label", None)
        roll["predicted_ordinal"] = predicted_ordinal
        roll["predicted_label"] = _ordinal_label("P", predicted_ordinal)
        roll["source_ordinal"] = source_ordinal
        roll["source_label"] = _ordinal_label("S", source_ordinal)
        roll.pop("roll_number", None)
    rolls[:] = ordered_all_rolls


def main() -> None:
    curator_doc = json.loads(CURATOR_ROLLS.read_text())
    outcomes_doc = json.loads(ROLL_OUTCOMES.read_text())
    evidence_doc = json.loads(EVIDENCE.read_text())
    obtained_doc = json.loads(OBTAINED_PERKS.read_text())
    directory = json.loads(DIRECTORY.read_text())["perks"]
    outstanding_doc = json.loads(OUTSTANDING.read_text())
    overrides_doc: dict = {}
    if ROLL_OVERRIDES.exists():
        overrides_doc = json.loads(ROLL_OVERRIDES.read_text())

    multi_overrides_doc = load_multi_grab_overrides()
    multi_overrides = multi_overrides_doc.get("chapter_roll_overrides") or {}
    obtained_lookup = _build_obtained_lookup(obtained_doc["perks"])
    (
        _early_source_occurrences_by_roll_number,
        _early_source_occurrences_by_chapter_index,
        early_source_occurrences_by_ordinal,
    ) = _source_roll_occurrences(curator_doc["rolls"], multi_overrides)
    early_source_occurrences_by_row_index = {
        int(info["source_row_index"]): info
        for info in early_source_occurrences_by_ordinal.values()
    }
    source_templates_by_source_ordinal = {
        int(info["source_ordinal"]): (
            int(info["source_row_index"]),
            curator_doc["rolls"][int(info["source_row_index"])],
        )
        for info in early_source_occurrences_by_ordinal.values()
    }
    cross_chapter_source_assignments: dict[int, str] = {}
    for mechanical_chapter, override in multi_overrides.items():
        for entry in override.get("rolls") or []:
            source_ordinal = entry.get("source_ordinal")
            if source_ordinal is None:
                continue
            try:
                source_ordinal = int(source_ordinal)
            except (TypeError, ValueError):
                continue
            template = source_templates_by_source_ordinal.get(source_ordinal)
            if template is None:
                continue
            _idx, source_row = template
            if str(source_row.get("chapter_num")) != str(mechanical_chapter):
                cross_chapter_source_assignments[source_ordinal] = str(mechanical_chapter)
    deferred_consumed_by_mention_chapter: dict[str, set[str]] = {}
    for mechanical_chapter, override in multi_overrides.items():
        for entry in override.get("rolls") or []:
            mention_chapter = _norm_chapter(
                entry.get("mention_chapter_num"), mechanical_chapter
            )
            if mention_chapter == str(mechanical_chapter):
                continue
            if (entry.get("outcome") or ("hit" if entry.get("perks") else "miss")) != "hit":
                continue
            for name in entry.get("perks") or []:
                deferred_consumed_by_mention_chapter.setdefault(
                    mention_chapter, set()
                ).add(_norm_name(name))

    curator_covered: set[str] = set()
    for row in curator_doc["rolls"]:
        if row.get("kind") in {"trigger", "roll", "miss"}:
            curator_covered.add(row["chapter_num"])

    curator_by_chapter: dict[str, list[tuple[int, dict]]] = {}
    for idx, row in enumerate(curator_doc["rolls"]):
        if row.get("kind") not in {"trigger", "roll", "miss"}:
            continue
        source_info = early_source_occurrences_by_row_index.get(idx)
        source_ordinal = (
            int(source_info["source_ordinal"])
            if source_info is not None else None
        )
        if (
            source_ordinal is not None
            and source_ordinal in cross_chapter_source_assignments
        ):
            continue
        deferred_names = deferred_consumed_by_mention_chapter.get(
            str(row.get("chapter_num")), set()
        )
        row_paid_names = {
            _norm_name(p.get("name"))
            for p in row.get("perks") or []
            if not p.get("free", False)
        }
        if row.get("kind") == "roll" and row_paid_names & deferred_names:
            continue
        curator_by_chapter.setdefault(row["chapter_num"], []).append((idx, row))
    for cn, override in multi_overrides.items():
        rows = curator_by_chapter.get(cn, [])
        if _override_needs_direct_rows(str(cn), override, rows):
            curator_by_chapter.setdefault(str(cn), rows)

    # ---- scheduler pass: per-chapter feasibility + slot assignment ----
    scheduler_inputs = _build_scheduler_inputs()
    for cn, curator_rows in curator_by_chapter.items():
        inp = scheduler_inputs.get(cn)
        if inp is None:
            continue
        non_trigger_count = sum(
            1 for _idx, row in curator_rows
            if row.get("kind") != "trigger"
        )
        extra = non_trigger_count - len(inp["slots"])
        if extra <= 0:
            continue
        words = int(inp["chapter_words"])
        existing = {int(slot.word_position) for slot in inp["slots"]}
        chapter_override = multi_overrides.get(cn) or {}
        override_rolls = chapter_override.get("rolls") or []
        for offset in range(extra):
            override_index = len(inp["predicted_slots"]) + offset
            pos = _explicit_extra_slot_position(
                override_rolls[override_index] if override_index < len(override_rolls) else None,
                words,
            )
            if pos is None:
                continue
            pos = max(0, min(words, int(pos)))
            while pos in existing and pos > 0:
                pos -= 1
            existing.add(pos)
            inp["slots"].append(SlotInput(
                word_position=pos,
                roll_trigger_cp_threshold=REGIMES[
                    regime_for_chapter(cn)
                ]["cp_per_roll"],
                source="curator_anchor",
            ))
        inp["slots"].sort(key=lambda slot: slot.word_position)
        inp["synthetic_slot_count"] = len(inp["slots"]) - len(inp["predicted_slots"])
    # Build a quick lookup of word offset (chapter-local) per slot index
    # per chapter, plus chapter_word_start for cumulative offsets.
    chapter_word_start_global = {
        cn: inp["chapter_word_start"] for cn, inp in scheduler_inputs.items()
    }
    (
        _source_occurrences_by_roll_number,
        source_occurrences_by_chapter_index,
        source_occurrences_by_ordinal,
    ) = _source_roll_occurrences(
        curator_doc["rolls"],
        multi_overrides,
        chapter_word_start_global,
    )
    source_templates_by_source_ordinal: dict[int, tuple[int, dict]] = {
        int(info["source_ordinal"]): (
            int(info["source_row_index"]),
            curator_doc["rolls"][int(info["source_row_index"])],
        )
        for info in source_occurrences_by_ordinal.values()
    }
    source_occurrences_by_row_index = {
        int(info["source_row_index"]): info
        for info in source_occurrences_by_ordinal.values()
    }
    scheduler_results: dict = {}
    strict_scheduler_results: dict = {}
    strict_infeasible_by_chapter: dict[str, dict] = {}
    infeasible_records: list[dict] = []
    for cn, inp in scheduler_inputs.items():
        # Apply chapter-level overrides (e.g., banked_cp_in nudge).
        ch_override = (overrides_doc.get("chapter_overrides") or {}).get(cn) or {}
        banked_in = (
            int(ch_override["banked_cp_in"])
            if "banked_cp_in" in ch_override
            else inp["banked_cp_in"]
        )
        strict_result = schedule_chapter(
            cn,
            inp["predicted_slots"],
            inp["hits"],
            banked_cp_in=banked_in,
            shadow_in=inp["shadow_in"],
            chapter_words=inp["chapter_words"],
            segments=inp.get("segments"),
        )
        if not strict_result.feasible:
            diag, expl = diagnose_infeasible(
                cn, inp["predicted_slots"], inp["hits"],
                banked_cp_in=banked_in,
                shadow_in=inp["shadow_in"],
                chapter_words=inp["chapter_words"],
                segments=inp.get("segments"),
            )
            strict_result.diagnostic = diag
            strict_result.explanation = expl
            strict_infeasible_by_chapter[cn] = {
                "chapter_num": cn,
                "banked_cp_in": banked_in,
                "predicted_slots": [
                    {"word_position": s.word_position,
                     "roll_trigger_cp_threshold": s.roll_trigger_cp_threshold,
                     "source": s.source}
                    for s in inp["predicted_slots"]
                ],
                "recorded_hits": [
                    {
                        "cost": h.cost,
                        "names": [
                            (p.get("perk_name") or p.get("name"))
                            for p in (h.paid_perks or ([h.perk] if h.perk else []))
                        ],
                    }
                    for h in inp["hits"]
                ],
                "diagnostic": diag,
                "explanation": expl,
            }
        strict_scheduler_results[cn] = strict_result

        result = schedule_chapter(
            cn,
            inp["slots"],
            inp["hits"],
            banked_cp_in=banked_in,
            shadow_in=inp["shadow_in"],
            chapter_words=inp["chapter_words"],
            segments=inp.get("segments"),
        )
        if not result.feasible:
            diag, expl = diagnose_infeasible(
                cn, inp["slots"], inp["hits"],
                banked_cp_in=banked_in,
                shadow_in=inp["shadow_in"],
                chapter_words=inp["chapter_words"],
                segments=inp.get("segments"),
            )
            result.diagnostic = diag
            result.explanation = expl
            # Curator-covered chapters with infeasible solver result are
            # NOT a hard pipeline error — curator is canonical there.
            # They surface as divergences instead.
            if cn not in curator_covered:
                infeasible_records.append({
                    "chapter_num": cn,
                    "banked_cp_in": banked_in,
                    "predicted_slots": [
                        {"word_position": s.word_position,
                         "roll_trigger_cp_threshold": s.roll_trigger_cp_threshold,
                         "source": s.source}
                        for s in inp["slots"]
                    ],
                    "recorded_hits": [
                        {
                            "cost": h.cost,
                            "names": [
                                (p.get("perk_name") or p.get("name"))
                                for p in (h.paid_perks or ([h.perk] if h.perk else []))
                            ],
                        }
                        for h in inp["hits"]
                    ],
                    "diagnostic": diag,
                    "explanation": expl,
                })
        scheduler_results[cn] = result

    match_idx = build_directory_lookup(directory)
    survey_designations = load_survey_designations()
    # Evidence is keyed by (chapter_num, slot_index) — NOT roll_number — to
    # tolerate drift between curator and predictor roll-number sequences. Each
    # predicted roll's evidence belongs to its mechanical chapter+slot; a
    # curator row joins to it via the same key derived from its matched
    # pred_slot.
    evidence_by_chapter_slot = {
        (str(e["chapter_num"]), int(e["slot_index"])): e
        for e in evidence_doc["rolls"]
    }
    outstanding_by_chapter = {
        c["chapter_num"]: c["before_chapter"]["by_constellation"]
        for c in outstanding_doc["chapters"]
    }

    outcome_by_chapter: dict[str, list[tuple[int, dict]]] = {}
    for idx, row in enumerate(outcomes_doc["rolls"]):
        outcome_by_chapter.setdefault(row["chapter_num"], []).append((idx, row))
    for rows in outcome_by_chapter.values():
        rows.sort(key=lambda item: (
            item[1].get("sequence_in_chapter") or 10**9,
            item[1].get("word_position")
            if item[1].get("word_position") is not None
            else 10**12,
            item[1].get("roll_number") or 10**9,
            item[0],
        ))

    all_chapters = sorted(
        set(curator_by_chapter) | set(outcome_by_chapter),
        key=sort_key_for_chapter,
    )

    rolls: list[dict] = []
    curator_solver_divergences: list[dict] = []
    for chapter_num in all_chapters:
        sched = scheduler_results.get(chapter_num)
        if chapter_num in curator_by_chapter:
            rows_raw = curator_by_chapter[chapter_num]
            chapter_override = multi_overrides.get(chapter_num)
            if chapter_override is not None:
                # Override beats curator: rebuild rows from override.
                if _override_needs_direct_rows(
                    chapter_num, chapter_override, rows_raw
                ):
                    synthetic = _direct_override_rows(
                        chapter_num,
                        rows_raw,
                        chapter_override,
                        obtained_lookup,
                        source_templates_by_source_ordinal,
                    )
                else:
                    synthetic = _restructure_curator_rows(
                        chapter_num, rows_raw, chapter_override,
                    )
                # Re-pack into (source_idx, row) tuples; source_idx points
                # at the original curator row (kept stable for evidence /
                # roll_key generation). Emit a unique roll_key per synthetic
                # row so override-split rows don't collide.
                rows = [
                    (s["_source_idx"], s) for s in synthetic
                ]
            else:
                rows = rows_raw
            total = len(rows)
            sched_inp = scheduler_inputs.get(chapter_num)
            pred_slot_by_local = {}
            if sched_inp is not None:
                pred_slot_by_local = {
                    max(
                        0,
                        int(pred["cp_offset"])
                        - int(sched_inp["chapter_word_start"]),
                    ): pred
                    for pred in sched_inp["predicted_rolls"]
                }
            non_trigger_seq = 0
            last_banked_after: int | None = None
            # Cross-check: compare curator outcomes vs scheduler decision.
            if sched and sched.feasible:
                curator_outcomes = [
                    "hit" if r[1]["kind"] != "miss" else "miss"
                    for r in rows
                    if r[1]["kind"] != "skipped"
                ]
                solver_outcomes = [a.outcome for a in sched.assignments]
                # Trim the solver's decision to the curator-recorded count
                # (curator may merge/duplicate roll_numbers; align by length
                # and emit a soft warning if they differ).
                if (
                    len(solver_outcomes) != len(curator_outcomes)
                    or solver_outcomes != curator_outcomes
                ):
                    curator_solver_divergences.append({
                        "chapter_num": chapter_num,
                        "curator_outcomes": curator_outcomes,
                        "solver_outcomes": solver_outcomes,
                        "solver_slack": sched.slack,
                    })
            for seq, (source_idx, row) in enumerate(rows, start=1):
                if row.get("kind") == "skipped":
                    if sched_inp is not None:
                        non_trigger_seq += 1
                    else:
                        non_trigger_seq += 1
                    continue
                roll_number = row.get("roll_number")
                # Evidence is joined later by (chapter_num, pred_slot.slot_index)
                # once the pred_slot is matched. Initialize lazily.
                ev: dict | None = None
                predicted_chapter_num: str | None = None
                outcome = "miss" if row["kind"] == "miss" else "hit"
                source_kind = row["kind"]
                available_cp = int_or_none(row.get("banked_before"))
                banked_after = int_or_none(row.get("banked_after"))
                paid_meta = None
                free_perks: list[dict] = []
                purchased_perks: list[dict] = []
                purchased_perk_cost_total: int | None = None
                purchased_perk_id = None
                purchased_perk_jump = None
                constellation = row.get("constellation")
                if outcome == "hit":
                    paid_meta, purchased_perks, free_perks = split_hit_perks(
                        match_idx,
                        row.get("perks") or [],
                        constellation,
                        survey_designations,
                    )
                    if paid_meta:
                        purchased_perk_id = paid_meta["id"]
                        purchased_perk_jump = paid_meta["jump"]
                        constellation = paid_meta["constellation"] or constellation
                    purchased_perk_cost_total = sum(
                        int(p["cost"]) for p in purchased_perks
                        if not p.get("free", False)
                    )
                ov_origin = row.get("_override_origin") if isinstance(row, dict) else None
                if ov_origin is not None:
                    roll_key = f"curator:{source_idx:04d}.{ov_origin}"
                else:
                    roll_key = f"curator:{source_idx:04d}"

                slot = None
                pred_slot = None
                assignment = None
                local_roll_index = None
                if row["kind"] != "trigger":
                    slot_idx = non_trigger_seq
                    local_roll_index = slot_idx + 1
                    non_trigger_seq += 1
                    if sched_inp is not None and slot_idx < len(sched_inp["slots"]):
                        slot = sched_inp["slots"][slot_idx]
                    if sched and sched.feasible and slot_idx < len(sched.assignments):
                        assignment = sched.assignments[slot_idx]
                    if slot is not None and slot.source == "predicted":
                        pred_slot = pred_slot_by_local.get(int(slot.word_position))
                if row.get("_override_direct") and assignment is not None:
                    if available_cp is None and row.get("_curator_added"):
                        available_cp = (
                            int(last_banked_after)
                            if last_banked_after is not None else 0
                        ) + int(assignment.roll_trigger_cp_threshold)
                    elif available_cp is None:
                        available_cp = assignment.available_cp
                    if banked_after is None and row.get("_curator_added"):
                        banked_after = available_cp
                    elif banked_after is None:
                        if outcome == "hit" and purchased_perk_cost_total is not None:
                            banked_after = max(
                                0,
                                int(available_cp) - int(purchased_perk_cost_total),
                            )
                        else:
                            banked_after = assignment.banked_cp_after_roll
                predicted_ordinal = (
                    int(pred_slot["roll_number"]) if pred_slot is not None else None
                )
                source_info = None
                if row.get("_source_ordinal") is not None:
                    source_info = source_occurrences_by_ordinal.get(
                        int(row["_source_ordinal"])
                    )
                if (
                    source_info is None
                    and row.get("_source_identity_inferred", True)
                ):
                    source_info = source_occurrences_by_row_index.get(int(source_idx))
                source_ordinal = (
                    int(source_info["source_ordinal"])
                    if source_info is not None else None
                )
                # Join evidence by (chapter, slot_index). pred_slot — which is
                # the chapter-local match for this curator row — is the
                # canonical source of both. Falls back to None when the row has
                # no matched predicted slot (rare; e.g. synthetic/trigger rows).
                if pred_slot is not None:
                    ev = evidence_by_chapter_slot.get(
                        (str(pred_slot["chapter_num"]), int(pred_slot["slot_index"]))
                    )
                predicted_chapter_num = (
                    ev.get("chapter_num") if ev else None
                )
                owner_chapter_num = (
                    row.get("_mention_chapter_num")
                    if row.get("_mention_chapter_num") is not None
                    else chapter_num
                )
                record = roll_base(
                    roll_key=roll_key,
                    roll_number=predicted_ordinal,
                    chapter_num=owner_chapter_num,
                    predicted_chapter_num=predicted_chapter_num,
                    source="curator_rolls",
                    source_kind=source_kind,
                    outcome=outcome,
                    source_row_index=source_idx,
                    ev=ev,
                )
                if pred_slot is not None:
                    # NB: cascade field is misnamed (CP-cumulative, not EPUB).
                    # Phase D drops it; keep stamping for now so chapter_facts
                    # consumers (TUI, etc.) keep working.
                    record["predicted_word_position_epub"] = int(pred_slot["cp_offset"])
                    # Canonical fields. epub_word_offset_curated defaults to
                    # predicted; a future manual-override mechanism (the
                    # "narrative-evidence override wins" rule) will overwrite
                    # only the _curated field.
                    record["epub_word_offset_predicted"] = int(pred_slot["epub_offset"])
                    record["epub_word_offset_curated"] = int(pred_slot["epub_offset"])
                word_position_local = int(slot.word_position) if slot is not None else None
                cum_word = (
                    chapter_word_start_global.get(chapter_num, 0) + word_position_local
                    if word_position_local is not None else None
                )
                policy = row.get("_display_position_policy")
                if policy is None:
                    policy = "mechanical"
                mention_word_position = row.get("_mention_word_position")
                if (
                    policy in {"mention", "source_marker"}
                    and mention_word_position is not None
                ):
                    display_chapter_num = owner_chapter_num
                    display_word_position = int(mention_word_position)
                    display_cum_word = (
                        chapter_word_start_global.get(display_chapter_num, 0)
                        + display_word_position
                    )
                elif policy == "section_end" and row["kind"] != "trigger":
                    display_chapter_num = owner_chapter_num
                    display_word_position = int(
                        sched_inp["chapter_words"]
                        if sched_inp is not None else 0
                    )
                    display_cum_word = (
                        chapter_word_start_global.get(display_chapter_num, 0)
                        + display_word_position
                    )
                elif policy == "section_start" and row["kind"] != "trigger":
                    display_chapter_num = owner_chapter_num
                    display_word_position = 0
                    display_cum_word = chapter_word_start_global.get(display_chapter_num, 0)
                else:
                    display_chapter_num = chapter_num
                    display_word_position = word_position_local
                    display_cum_word = cum_word
                miss_estimate = (
                    miss_cost_estimate(
                        outstanding_by_chapter,
                        chapter_num,
                        constellation,
                        available_cp,
                    )
                    if outcome == "miss"
                    else None
                )
                principal_name = paid_meta["name"] if paid_meta else None
                principal_instance = (
                    paid_meta.get("instance") if paid_meta else None
                )
                source_chapter_num = (
                    str(source_info["source_chapter_num"])
                    if source_info is not None else None
                )
                source_chapter_ordinal = (
                    int(source_info["source_chapter_ordinal"])
                    if source_info is not None else None
                )
                source_roll_label = (
                    str(source_info["source_roll_label"])
                    if source_info is not None else None
                )
                source_word_position = (
                    source_info.get("source_word_position")
                    if source_info is not None else None
                )
                source_cumulative_word_offset = (
                    source_info.get("source_cumulative_word_offset")
                    if source_info is not None else None
                )
                visible_chapters = _visible_chapter_nums(
                    owner_chapter_num,
                    chapter_num,
                    display_chapter_num,
                    source_chapter_num,
                )
                record.update({
                    "mechanical_chapter_num": chapter_num,
                    "mechanical_word_position": word_position_local,
                    "mechanical_cumulative_word_offset": (
                        int(cum_word) if cum_word is not None else None
                    ),
                    "mention_chapter_num": owner_chapter_num,
                    "mention_word_position": (
                        int(mention_word_position)
                        if mention_word_position is not None else None
                    ),
                    "display_position_policy": policy,
                    "display_chapter_num": display_chapter_num,
                    "display_word_position": display_word_position,
                    "display_cumulative_word_offset": (
                        int(display_cum_word) if display_cum_word is not None else None
                    ),
                    "source_chapter_num": source_chapter_num,
                    "source_chapter_ordinal": source_chapter_ordinal,
                    "source_roll_label": source_roll_label,
                    "source_word_position": (
                        int(source_word_position)
                        if source_word_position is not None else None
                    ),
                    "source_cumulative_word_offset": (
                        int(source_cumulative_word_offset)
                        if source_cumulative_word_offset is not None else None
                    ),
                    "visible_chapter_nums": visible_chapters,
                    "constellation": constellation,
                    "constellation_revealed": bool(row.get("constellation_revealed", False)),
                    "available_cp": available_cp,
                    "banked_cp_after_roll": banked_after,
                    "purchased_perk_id": purchased_perk_id,
                    "purchased_perks": purchased_perks,
                    "purchased_perk_cost_total": purchased_perk_cost_total,
                    "purchased_perk_jump": purchased_perk_jump,
                    "free_perks": free_perks,
                    "rolled_perk_name": principal_name,
                    "rolled_perk_instance": principal_instance,
                    "rolled_perk_cost": (
                        purchased_perk_cost_total if outcome == "hit" else miss_estimate
                    ),
                    "miss_cost_estimate": miss_estimate,
                    "evidence_quotes": row.get("_evidence_quotes") or [],
                    "raw": row.get("raw"),
                    "roll_sequence_in_chapter": seq,
                    "rolls_in_chapter": total,
                    "slot_source": (
                        "curator+override" if ov_origin is not None
                        else "curator"
                    ),
                    "word_position": display_word_position,
                    "cumulative_word_offset": (
                        int(display_cum_word) if display_cum_word is not None else None
                    ),
                })
                _record_roll_identity(
                    record,
                    predicted_ordinal=predicted_ordinal,
                    source_ordinal=source_ordinal,
                    association_source=_association_source_for(
                        predicted_ordinal=predicted_ordinal,
                        source_ordinal=source_ordinal,
                        curated=bool(
                            chapter_override is not None
                            or ov_origin is not None
                            or row.get("_source_ordinal") is not None
                            or row.get("_override_direct")
                        ),
                    ),
                )
                if row.get("_curator_added"):
                    record["curator_added"] = True
                rolls.append(record)
                if banked_after is not None:
                    last_banked_after = int(banked_after)
            continue

        rows = outcome_by_chapter.get(chapter_num, [])
        total = len(rows)
        sched_inp = scheduler_inputs.get(chapter_num)
        pred_slot_by_local = {}
        if sched_inp is not None:
            pred_slot_by_local = {
                max(
                    0,
                    int(pred["cp_offset"])
                    - int(sched_inp["chapter_word_start"]),
                ): pred
                for pred in sched_inp["predicted_rolls"]
            }
        sched_assignments = (
            sched.assignments if sched and sched.feasible else None
        )
        chapter_override = multi_overrides.get(chapter_num) or {}
        fallback_override_rolls = chapter_override.get("rolls") or []
        # Build a fast lookup: which "hit_index" (in the scheduler's hit
        # list, which is paid-acquisitions in epub_sequence) corresponds
        # to each slot.
        for seq, (source_idx, row) in enumerate(rows, start=1):
            manual_entry = _fallback_projection_override(
                fallback_override_rolls[seq - 1]
                if seq - 1 < len(fallback_override_rolls) else None
            )
            roll_number = row.get("roll_number")
            # Evidence is joined later by (chapter, slot_index) using the
            # matched pred_slot. Initialize lazily.
            ev: dict | None = None
            assignment = (
                sched_assignments[seq - 1]
                if sched_assignments and seq - 1 < len(sched_assignments)
                else None
            )
            if assignment is not None:
                outcome = assignment.outcome
            else:
                outcome = "hit" if row.get("outcome") == "hit" else "miss"
            if (
                manual_entry is not None
                and manual_entry.get("outcome") in {"hit", "miss"}
            ):
                outcome = str(manual_entry["outcome"])
            # Pick perk source:
            # - If the scheduler assigned a hit, use the paid_perks list
            #   from sched_inp.hits[assignment.hit_index].
            # - Otherwise (no scheduler / miss), keep the row's perk
            #   metadata if it exists (so manual data is preserved).
            raw_paid_list: list[dict] = []
            free_perks_payload: list[dict] = []
            assigned_hit = None
            if (
                outcome == "hit"
                and assignment is not None
                and assignment.hit_index is not None
                and sched_inp is not None
            ):
                assigned_hit = sched_inp["hits"][assignment.hit_index]
                raw_paid_list = list(
                    assigned_hit.paid_perks
                    or ([assigned_hit.perk] if assigned_hit.perk else [])
                )
                free_perks_payload = assigned_hit.free_perks or []
            elif outcome == "hit":
                row_paid = row.get("paid_perks") or []
                if row_paid:
                    raw_paid_list = list(row_paid)
                elif row.get("perk"):
                    raw_paid_list = [row["perk"]]
                free_perks_payload = row.get("free_perks") or []
            principal_raw = raw_paid_list[0] if raw_paid_list else None
            paid_meta = (
                perk_meta(
                    match_idx, principal_raw,
                    principal_raw.get("constellation") if principal_raw else None,
                )
                if principal_raw
                else None
            )
            purchased_perks: list[dict] = []
            for p in raw_paid_list:
                pm = perk_meta(match_idx, p, p.get("constellation"))
                purchased_perks.append(
                    _purchased_perk_record(pm, p, survey_designations)
                )
            purchased_perk_cost_total = (
                sum(int(p["cost"]) for p in purchased_perks)
                if purchased_perks else None
            )
            free_perks = []
            if paid_meta:
                for raw_free in free_perks_payload:
                    # raw_free dicts come from obtained_perks (perk_name) or
                    # roll_outcomes (name); perk_meta handles both.
                    free_meta = perk_meta(
                        match_idx,
                        raw_free,
                        paid_meta["constellation"],
                    )
                    free_perks.append({
                        "id": free_meta["id"],
                        "name": free_meta["name"],
                        "jump": free_meta["jump"],
                        "constellation": free_meta["constellation"],
                    })
                    designation = survey_designation_for(
                        survey_designations,
                        free_meta["name"],
                        free_meta["constellation"],
                    )
                    if designation:
                        free_perks[-1]["survey_designation"] = designation
            if assignment is not None:
                available_cp = assignment.available_cp
                banked_cp_after = assignment.banked_cp_after_roll
            else:
                available_cp = int_or_none(row.get("available_cp"))
                if available_cp is None:
                    available_cp = int_or_none(row.get("roll_trigger_cp_threshold"))
                banked_cp_after = int_or_none(row.get("banked_cp_after_roll"))
            manual_constellation = (
                manual_entry.get("constellation")
                if manual_entry is not None
                and manual_entry.get("constellation") is not None
                else None
            )
            constellation = (
                paid_meta["constellation"] if paid_meta else manual_constellation
            )
            miss_estimate = (
                miss_cost_estimate(
                    outstanding_by_chapter,
                    chapter_num,
                    constellation,
                    available_cp,
                )
                if outcome == "miss"
                else None
            )
            owner_chapter_num = (
                assigned_hit.mention_chapter_num
                if assigned_hit is not None and assigned_hit.mention_chapter_num
                else chapter_num
            )
            if (
                manual_entry is not None
                and manual_entry.get("mention_chapter_num") is not None
            ):
                owner_chapter_num = _norm_chapter(
                    manual_entry.get("mention_chapter_num"), chapter_num
                )
            record = roll_base(
                roll_key=f"interpolated:{source_idx:04d}",
                roll_number=roll_number,
                chapter_num=owner_chapter_num,
                predicted_chapter_num=ev.get("chapter_num") if ev else chapter_num,
                source="roll_outcomes",
                source_kind="interpolated",
                outcome=outcome,
                source_row_index=source_idx,
                ev=ev,
            )
            if record["predicted_word_position_epub"] is None:
                record["predicted_word_position_epub"] = row.get("word_position")
            record["evidence_kind"] = (
                "synthetic" if row.get("source") == "synthetic"
                else record["evidence_kind"]
            )
            ch_start = chapter_word_start_global.get(chapter_num, 0)
            # Prefer the scheduler's slot word_position (uniformly
            # chapter-local). Fall back to the row's word_position only
            # if no scheduler input exists; treat that as chapter-local
            # too if it's clearly small (synthetic) and as cumulative
            # otherwise.
            if sched_inp is not None and seq - 1 < len(sched_inp["slots"]):
                slot = sched_inp["slots"][seq - 1]
                word_position_local = slot.word_position
                cum_word = ch_start + word_position_local
                pred_slot = (
                    pred_slot_by_local.get(int(slot.word_position))
                    if slot.source == "predicted" else None
                )
                if pred_slot is not None:
                    ev2 = evidence_by_chapter_slot.get(
                        (str(pred_slot["chapter_num"]),
                         int(pred_slot["slot_index"]))
                    )
                    record["roll_number"] = int(pred_slot["roll_number"])
                    record["predicted_chapter_num"] = (
                        ev2.get("chapter_num") if ev2 else pred_slot["chapter_num"]
                    )
                    record["chapter_attribution_disagreement"] = (
                        record["predicted_chapter_num"] is not None
                        and str(record["predicted_chapter_num"]) != str(record["chapter_num"])
                    )
                    # NB: cascade field is CP-cumulative; Phase D drops it.
                    record["predicted_word_position_epub"] = int(pred_slot["cp_offset"])
                    record["epub_word_offset_predicted"] = int(pred_slot["epub_offset"])
                    record["epub_word_offset_curated"] = int(pred_slot["epub_offset"])
                    record["predicted_char_offset_in_chapter"] = (
                        ev2.get("predicted_char_offset") if ev2 else None
                    )
                    record["anchor_char_offset_in_chapter"] = anchor_offset(ev2)
                    record["evidence_kind"] = normalized_evidence_kind(
                        ev2,
                        "interpolated",
                    )
                else:
                    record["roll_number"] = None
                    record["predicted_chapter_num"] = chapter_num
                    record["chapter_attribution_disagreement"] = False
                    record["predicted_word_position_epub"] = None
                    record["epub_word_offset_predicted"] = None
                    record["epub_word_offset_curated"] = None
                    record["predicted_char_offset_in_chapter"] = None
                    record["anchor_char_offset_in_chapter"] = None
                    record["evidence_kind"] = "synthetic"
            else:
                row_wp = row.get("word_position")
                if row_wp is None:
                    word_position_local = None
                    cum_word = None
                elif int(row_wp) >= ch_start:
                    cum_word = int(row_wp)
                    word_position_local = cum_word - ch_start
                else:
                    word_position_local = int(row_wp)
                    cum_word = ch_start + word_position_local
            policy = (
                assigned_hit.display_position_policy
                if assigned_hit is not None else "mechanical"
            )
            mention_word_position = (
                assigned_hit.mention_word_position
                if assigned_hit is not None else None
            )
            evidence_quotes = (
                assigned_hit.evidence_quotes
                if assigned_hit is not None else []
            )
            constellation_revealed = False
            if manual_entry is not None:
                if manual_entry.get("display_position_policy") is not None:
                    policy = manual_entry.get("display_position_policy")
                if manual_entry.get("mention_word_position") is not None:
                    mention_word_position = manual_entry.get("mention_word_position")
                evidence_quotes = _evidence_quotes(manual_entry)
                if manual_entry.get("constellation") is not None:
                    constellation_revealed = True
            if (
                policy in {"mention", "source_marker"}
                and mention_word_position is not None
            ):
                display_chapter_num = owner_chapter_num
                display_word_position = int(mention_word_position)
                display_cum_word = (
                    chapter_word_start_global.get(display_chapter_num, 0)
                    + display_word_position
                )
            elif policy == "section_end" and record["source_kind"] != "trigger":
                display_chapter_num = owner_chapter_num
                display_word_position = int(
                    sched_inp["chapter_words"] if sched_inp is not None else 0
                )
                display_cum_word = (
                    chapter_word_start_global.get(display_chapter_num, 0)
                    + display_word_position
                )
            elif policy == "section_start" and record["source_kind"] != "trigger":
                display_chapter_num = owner_chapter_num
                display_word_position = 0
                display_cum_word = chapter_word_start_global.get(display_chapter_num, 0)
            else:
                display_chapter_num = chapter_num
                display_word_position = word_position_local
                display_cum_word = cum_word
            visible_chapters = _visible_chapter_nums(
                owner_chapter_num,
                chapter_num,
                display_chapter_num,
            )
            record.update({
                "mechanical_chapter_num": chapter_num,
                "mechanical_word_position": word_position_local,
                "mechanical_cumulative_word_offset": (
                    int(cum_word) if cum_word is not None else None
                ),
                "mention_chapter_num": owner_chapter_num,
                "mention_word_position": (
                    int(mention_word_position)
                    if mention_word_position is not None else None
                ),
                "display_position_policy": policy,
                "display_chapter_num": display_chapter_num,
                "display_word_position": display_word_position,
                "display_cumulative_word_offset": (
                    int(display_cum_word) if display_cum_word is not None else None
                ),
                "source_chapter_num": None,
                "source_chapter_ordinal": None,
                "source_roll_label": None,
                "source_word_position": None,
                "source_cumulative_word_offset": None,
                "visible_chapter_nums": visible_chapters,
                "constellation": constellation,
                "constellation_revealed": constellation_revealed,
                "available_cp": available_cp,
                "banked_cp_after_roll": banked_cp_after,
                "purchased_perk_id": paid_meta["id"] if paid_meta else None,
                "purchased_perks": purchased_perks,
                "purchased_perk_cost_total": purchased_perk_cost_total,
                "purchased_perk_jump": paid_meta["jump"] if paid_meta else None,
                "free_perks": free_perks,
                "rolled_perk_name": paid_meta["name"] if paid_meta else None,
                "rolled_perk_instance": (
                    paid_meta.get("instance") if paid_meta else None
                ),
                "rolled_perk_cost": (
                    purchased_perk_cost_total if paid_meta else miss_estimate
                ),
                "miss_cost_estimate": miss_estimate,
                "evidence_quotes": evidence_quotes,
                "raw": None,
                "roll_sequence_in_chapter": seq,
                "rolls_in_chapter": total,
                "slot_source": "solver" if assignment is not None else "interpolated",
                "word_position": display_word_position,
                "cumulative_word_offset": display_cum_word,
            })
            predicted_ordinal = int_or_none(record.get("roll_number"))
            _record_roll_identity(
                record,
                predicted_ordinal=predicted_ordinal,
                source_ordinal=None,
                association_source=_association_source_for(
                    predicted_ordinal=predicted_ordinal,
                    source_ordinal=None,
                    curated=False,
                ),
            )
            rolls.append(record)

    # ---- apply overrides as a final pass ----
    overrides_applied = _apply_overrides(rolls, overrides_doc)

    chapter_order = [
        c["chapter_num"]
        for c in sorted(
            json.loads(CHAPTERS_JSON.read_text())["chapters"],
            key=lambda c: tuple(c["sort_key"]),
        )
    ]
    _stamp_explicit_ordinals(rolls, chapter_order)
    (
        hit_accounting_corrections,
        hit_accounting_missing_available_cp,
    ) = _normalize_hit_accounting(rolls)
    cp_ledger_checkpoint_applications = _apply_cp_ledger_checkpoints(
        rolls,
        chapter_order=chapter_order,
        scheduler_inputs=scheduler_inputs,
        outstanding_by_chapter=outstanding_by_chapter,
    )

    known_attempt_count_by_chapter: dict[str, int] = {}
    for roll in rolls:
        if roll.get("source_kind") == "trigger":
            continue
        cn = str(roll.get("mechanical_chapter_num") or roll.get("chapter_num"))
        known_attempt_count_by_chapter[cn] = (
            known_attempt_count_by_chapter.get(cn, 0) + 1
        )

    payload = {
        "schema_version": 1,
        "_source": (
            "Merged from data/derived/rolls.json and "
            "data/derived/roll_outcomes.json with perk/evidence enrichment."
        ),
        "_method": (
            "Use every trigger/roll/miss row from the curator roll log as "
            "authoritative for chapters it covers, preserving source order and "
            "duplicate roll numbers. For chapters without curator rows, the "
            "scheduler (scripts/roll_scheduler.py) decides which predicted "
            "slots are hits via latest-feasible backtracking. Free perks "
            "attach to the paid hit and do not consume separate roll slots. "
            "Manual roll/chapter overrides land last (slot_source='override'). "
            "After final roll ordering, hit banked-CP-after values are "
            "normalized from available CP minus paid perk costs. Perk costs "
            "are already normalized to effective Forge CP; cost_unit is "
            "provenance metadata. Quote-carried CP ledger checkpoints apply "
            "after this normalization and reset downstream derived ledger "
            "balances without moving predicted slots."
        ),
        "_caveat": (
            "Curator rows are the trusted source. Solver rows are "
            "constraint-satisfaction reconstructions; ambiguous chapters "
            "and infeasible chapters are listed in roll_validation.json."
        ),
        "_counts": {
            "rolls_emitted": len(rolls),
            "curator_rows": sum(1 for r in rolls if r["source"] == "curator_rolls"),
            "interpolated_rows": sum(1 for r in rolls if r["source"] == "roll_outcomes"),
            "hits": sum(1 for r in rolls if r["outcome"] == "hit"),
            "misses": sum(1 for r in rolls if r["outcome"] == "miss"),
            "triggers": sum(1 for r in rolls if r["source_kind"] == "trigger"),
            "free_perks": sum(len(r["free_perks"]) for r in rolls),
        },
        "rolls": rolls,
    }

    write_validated_json(OUT, payload, "roll_facts")

    # ---- write validation report ----
    ambiguous = [
        {
            "chapter_num": cn,
            "feasible_assignment_count": r.ambiguity,
            "slack": r.slack,
        }
        for cn, r in scheduler_results.items()
        if r.feasible and r.ambiguity > 1
    ]
    ambiguous_by_chapter = {
        row["chapter_num"]: row for row in ambiguous
    }
    chapter_checks: list[dict] = []
    for cn in chapter_order:
        inp = scheduler_inputs.get(cn, {})
        predicted_count = len(inp.get("predicted_slots") or [])
        required_paid_count = len(inp.get("hits") or [])
        known_attempt_count = known_attempt_count_by_chapter.get(cn, 0)
        issues: list[dict] = []
        issues.extend(_manual_override_issues(cn, multi_overrides.get(cn)))

        if required_paid_count > predicted_count:
            issues.append({
                "code": "paid_rolls_exceed_predicted_slots",
                "severity": "error",
                "message": (
                    f"{required_paid_count} paid roll unit(s) require "
                    f"more slots than {predicted_count} predicted roll(s)."
                ),
            })
        if known_attempt_count > predicted_count:
            issues.append({
                "code": "known_attempts_exceed_predicted_slots",
                "severity": "error",
                "message": (
                    f"{known_attempt_count} known roll attempt(s) require "
                    f"more slots than {predicted_count} predicted roll(s)."
                ),
            })

        strict_result = strict_scheduler_results.get(cn)
        cost_schedule_ok = True
        if strict_result is not None and not strict_result.feasible:
            cost_schedule_ok = False
            issues.append({
                "code": "cost_schedule_infeasible",
                "severity": "error",
                "message": strict_result.explanation,
                "diagnostic": strict_result.diagnostic,
            })
        if cn in ambiguous_by_chapter:
            issues.append({
                "code": "ambiguous_schedule",
                "severity": "info",
                "message": (
                    f"{ambiguous_by_chapter[cn]['feasible_assignment_count']} "
                    "feasible hit/slot assignments."
                ),
            })

        issues, resolved_issue_codes = _apply_issue_resolutions(
            issues, multi_overrides.get(cn)
        )
        status, has_discrepancy, raw_has_discrepancy = _validation_status(issues)
        if cn in multi_overrides and _has_structural_roll_override(multi_overrides[cn]):
            source_priority = "vetted_curated"
        elif cn in curator_by_chapter:
            source_priority = "curator_log"
        elif required_paid_count:
            source_priority = "derived_perk_list"
        else:
            source_priority = "none"

        chapter_checks.append({
            "chapter_num": cn,
            "status": status,
            "has_discrepancy": has_discrepancy,
            "raw_has_discrepancy": raw_has_discrepancy,
            "resolved_issue_codes": resolved_issue_codes,
            "source_priority": source_priority,
            "predicted_roll_count": predicted_count,
            "required_paid_roll_count": required_paid_count,
            "known_attempt_count": known_attempt_count,
            "paid_roll_capacity_ok": required_paid_count <= predicted_count,
            "known_attempt_capacity_ok": known_attempt_count <= predicted_count,
            "cost_schedule_ok": cost_schedule_ok,
            "synthetic_slot_count": int(inp.get("synthetic_slot_count") or 0),
            "issues": issues,
        })

    strict_infeasible_records = [
        strict_infeasible_by_chapter[cn]
        for cn in chapter_order
        if cn in strict_infeasible_by_chapter
    ]
    validation_payload = {
        "_generated": _dt.datetime.now(_dt.UTC).isoformat(timespec="seconds"),
        "_source": "scripts/derive_roll_facts.py",
        "summary": {
            "chapters_total": len(scheduler_results),
            "feasible": sum(1 for r in strict_scheduler_results.values() if r.feasible),
            "infeasible": len(strict_infeasible_records),
            "model_discrepancies": sum(
                1 for row in chapter_checks if row["has_discrepancy"]
            ),
            "curator_divergences": len(curator_solver_divergences),
            "ambiguous": len(ambiguous),
            "overrides_applied": overrides_applied,
        },
        "chapter_checks": chapter_checks,
        "infeasible": strict_infeasible_records,
        "generation_infeasible": infeasible_records,
        "curator_solver_divergences": curator_solver_divergences,
        "ambiguous_chapters": ambiguous,
        "hit_accounting_corrections": hit_accounting_corrections,
        "hit_accounting_missing_available_cp": (
            hit_accounting_missing_available_cp
        ),
        "cp_ledger_checkpoint_applications": cp_ledger_checkpoint_applications,
    }
    VALIDATION_OUT.parent.mkdir(parents=True, exist_ok=True)
    VALIDATION_OUT.write_text(
        json.dumps(validation_payload, indent=2, ensure_ascii=False) + "\n"
    )

    counts = payload["_counts"]
    print(f"wrote {OUT.relative_to(ROOT)}: {counts['rolls_emitted']} rows")
    print(
        f"  curator: {counts['curator_rows']}, "
        f"interpolated: {counts['interpolated_rows']}"
    )
    print(
        f"  hits: {counts['hits']}, misses: {counts['misses']}, "
        f"free perks: {counts['free_perks']}"
    )
    print(f"wrote {VALIDATION_OUT.relative_to(ROOT)}: "
          f"infeasible={len(strict_infeasible_records)}, "
          f"model_discrepancies={validation_payload['summary']['model_discrepancies']}, "
          f"divergences={len(curator_solver_divergences)}, "
          f"ambiguous={len(ambiguous)}, overrides={overrides_applied}")
    if strict_infeasible_records:
        print("  INFEASIBLE chapters:")
        for r in strict_infeasible_records:
            print(f"    ch{r['chapter_num']}: {r['diagnostic']} - {r['explanation']}")


def derive_chapter_facts(chapter_num: str, *, source_path: Path | None = None) -> list[dict]:
    """Return the list of roll-facts rows for a single chapter.

    Phase 0 scaffold for Forge Curator's per-chapter recompute. The
    canonical implementation reads the already-derived
    ``roll_facts.json`` and slices it; Phase 3 will replace this with
    a true in-memory recompute that runs the predict/outcomes/facts
    pipeline for the chapter only and merges into the stable upstream
    state.
    """
    p = source_path or OUT
    if not p.exists():
        raise FileNotFoundError(
            f"derive_chapter_facts: {p} missing — run derive_roll_facts.py first"
        )
    doc = json.loads(p.read_text())
    cn = str(chapter_num)
    return [r for r in doc.get("rolls", []) if str(r.get("chapter_num")) == cn]


if __name__ == "__main__":
    main()
