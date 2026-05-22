"""Build the canonical perk directory.

Source:  data/raw/Brocktons_Celestial_Forge_Reference.xlsx#Unabridged List

The Unabridged List is the master roster of every perk in the gacha:
constellation, name, jump (source media), and a status flag. Multi-perk
rows have newline-separated names in column 2; we expand them so each
perk gets its own row.

We enrich each perk with:
  - cost / cost_text / repeatable / description from perks_catalog.json
    (the Complete List of Perks sheet) when we can match
  - acquisition info (chapter_num, epub_sequence) from obtained_perks.json
    when matched. Repeatable perks use their EARLIEST acquisition chapter.

Filter: drop only Status == "Excluded"; keep Obtained, Available,
Partial, Locked, Unknown, Repeatable.

Output: data/derived/perk_directory.json (validated against
data/derived/_schemas/perk_directory.schema.json).
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

from _common import write_validated_json
from perk_name_resolver import (
    JUMP_ALIASES,
    _name_prefix_variants,
    _normalize,
    _normalized_word_prefixes,
    build_alias_lookup,
    load_perk_aliases,
)

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Brocktons_Celestial_Forge_Reference.xlsx"
PERKS_CATALOG_JSON = ROOT / "data" / "derived" / "perks_catalog.json"
OBTAINED_JSON = ROOT / "data" / "derived" / "obtained_perks.json"
UNABRIDGED_OVERLAY_JSON = (
    ROOT / "data" / "manual" / "unabridged_overlay.json"
)
OUT = ROOT / "data" / "derived" / "perk_directory.json"


_CONSTELLATIONS = {
    "Toolkits", "Knowledge", "Vehicles", "Time", "Crafting",
    "Clothing", "Magic", "Quality", "Size",
    "Resources and Durability", "Magitech", "Alchemy",
    "Capstone", "Personal Reality", "Felyne Perks",
}

_KEEP_STATUSES = {
    "Obtained", "Available", "Partial", "Locked", "Unknown", "Repeatable",
}


def _id_for(constellation: str, jump: str, name: str) -> str:
    return f"{constellation}__{jump}__{_normalize(name)}"


def _norm_chapter(ch: str) -> tuple[int, int]:
    """Sort key for chapter ids like '12' or '43.1'."""
    if not ch:
        return (0, 0)
    if "." in ch:
        a, b = ch.split(".", 1)
        return (int(a), int(b))
    return (int(ch), 0)


@dataclass
class DirectoryPerk:
    id: str
    constellation: str
    name: str
    jump: str
    status: str
    cost: int | None
    cost_text: str | None
    cost_unit: str | None
    free: bool
    repeatable: bool
    description: str
    acquired_chapter_num: str | None
    acquired_epub_sequence: int | None
    first_acquired_at_word_offset: int | None
    matched_to_obtained: bool
    # "unabridged" for entries derived from the Unabridged List (the
    # canonical roster of rollable perks); "obtained_supplemental" for
    # acquisitions that don't match any Unabridged row but Joe still
    # got them (typically curator added them ad-hoc post-roll).
    source: str
    # Sub-instance acquisitions folded into this entry, in chapter
    # order. Empty for non-repeatable perks. Each entry is {name,
    # chapter_num, epub_sequence}. The parent's acquired_chapter_num
    # is the earliest of these.
    acquired_instances: list[dict]


# ---------- catalog index ---------------------------------------------------


def _build_catalog_index(catalog: dict) -> dict[str, dict]:
    """Build progressive lookup tables from perks_catalog.json.

    Returned: {
      "exact": (name, source) -> catalog perk,
      "norm":  (norm_name, norm_source) -> catalog perk,
    }
    """
    by_exact: dict[tuple[str, str], dict] = {}
    by_norm: dict[tuple[str, str], dict] = {}
    for p in catalog.get("perks", []):
        name = p.get("name", "")
        source = p.get("source", "") or ""
        by_exact.setdefault((name, source), p)
        by_norm.setdefault((_normalize(name), _normalize(source)), p)
    return {"exact": by_exact, "norm": by_norm}


def _catalog_lookup(idx: dict, name: str, jump: str) -> dict | None:
    """Look up cost/description for a directory perk.

    Tries exact (name, jump), then case+punct-normalized (name, jump).
    No name-only fallback: cost is jump-specific (Workshop costs 100
    in Personal Reality but 200 in Samurai Jack), so folding by name
    alone would silently use the wrong cost.
    """
    if (name, jump) in idx["exact"]:
        return idx["exact"][(name, jump)]
    key = (_normalize(name), _normalize(jump))
    if key in idx["norm"]:
        return idx["norm"][key]
    return None


# ---------- obtained index --------------------------------------------------


def _build_obtained_index(
    obtained: dict,
    aliases: dict[str, list[str]],
    alias_lookup: dict[str, str],
) -> dict[str, dict]:
    """Index obtained_perks acquisitions for join.

    Repeatable perks may appear multiple times; collect *all* matches
    per key so we can pick the earliest chapter.

    Each obtained entry registers under multiple keys to give the
    directory's lookup a fighting chance:

      - exact (name, jump) and case+punct-normalized (name, jump)
      - separator-split prefix AND suffix variants (e.g. "Workshop:
        Metalworking" -> "Workshop", "Innate Talent: Alchemist" ->
        "Alchemist")
      - normalized word-prefixes (e.g. "Minor Blessing Aphrodite -
        Beauty" -> "minor blessing", which matches the directory's
        repeatable "Minor Blessing" parent in Percy Jackson)
      - any JUMP_ALIASES form: a "Star Trek - TNG+DS9" entry also
        registers under "Star Trek: TNG" so the directory's TNG entry
        finds it
      - canonical alias from data/manual/perk_aliases.json: typos
        like "POWER OVERHELMING" / "Seeds and Seedlings" register
        under their canonical names; sub-instance aliases like
        "Workshop: Electronics" / "Thaumaturgical Focus: Alchemy"
        register under the parent canonical too

    Same-jump scoping prevents cross-jump pollution (Samurai Jack's
    "Workshop" never resolves to Personal Reality Workshops).
    """
    by_exact: dict[tuple[str, str], list[dict]] = defaultdict(list)
    by_norm: dict[tuple[str, str], list[dict]] = defaultdict(list)
    by_name_only: dict[str, list[dict]] = defaultdict(list)

    def _register(p: dict, name: str, jump: str) -> None:
        nn = _normalize(name)
        nj = _normalize(jump)
        by_exact[(name, jump)].append(p)
        by_norm[(nn, nj)].append(p)
        by_name_only[nn].append(p)
        for variant in _name_prefix_variants(name):
            vn = _normalize(variant)
            by_norm[(vn, nj)].append(p)
            by_name_only[vn].append(p)
        for nv in _normalized_word_prefixes(name):
            by_norm[(nv, nj)].append(p)
            by_name_only[nv].append(p)

    for p in obtained.get("perks", []):
        name = p.get("perk_name", "")
        jump = p.get("jump") or ""

        # All name forms to register this obtained perk under:
        # - the curator-typed name itself
        # - if it's an alias, the canonical
        # - if it's a canonical, all its aliases (covers the inverse
        #   case where the Unabridged xlsx has the alias spelling
        #   and the curator log has the canonical)
        name_forms: list[str] = [name]
        canonical = alias_lookup.get(name)
        if canonical and canonical != name:
            name_forms.append(canonical)
        for alias_form in aliases.get(name, []):
            if alias_form != name:
                name_forms.append(alias_form)

        for form in name_forms:
            _register(p, form, jump)

        # Jump alias: also register under the canonical jump so the
        # directory entry (which uses the canonical jump) can find it.
        aliased_jump = JUMP_ALIASES.get(jump)
        if aliased_jump and aliased_jump != jump:
            for form in name_forms:
                _register(p, form, aliased_jump)

    return {"exact": by_exact, "norm": by_norm, "name_only": by_name_only}


def _obtained_lookup(idx: dict, name: str, jump: str) -> list[dict]:
    """Return acquisitions matching this directory entry, deduping by
    object identity. Combines:

      1. exact (name, jump)
      2. case+punct-normalized (name, jump) - includes
         prefix/suffix/word-prefix variants registered at index time
      3. case+punct-normalized name only - only when the directory
         entry has no jump (currently never; kept for safety)

    A single obtained acquisition can register under several keys, so
    we dedupe by id() to avoid double-counting in the spot-check.
    """
    seen: set[int] = set()
    out: list[dict] = []

    def _extend(entries: list[dict]) -> None:
        for p in entries:
            pid = id(p)
            if pid in seen:
                continue
            seen.add(pid)
            out.append(p)

    _extend(idx["exact"].get((name, jump), []))
    _extend(idx["norm"].get((_normalize(name), _normalize(jump)), []))
    if not jump:
        _extend(idx["name_only"].get(_normalize(name), []))
    return out


def _earliest(acquisitions: list[dict]) -> dict:
    """The acquisition with the lowest chapter; ties broken by epub_sequence."""
    return min(
        acquisitions,
        key=lambda p: (
            _norm_chapter(p.get("chapter_num", "0")),
            p.get("epub_sequence", 0),
        ),
    )


# ---------- main ------------------------------------------------------------


def _norm_cell(value) -> str:
    return str(value).replace("\r", "").strip() if value is not None else ""


def _normalize_constellation(raw: str) -> str:
    """Trim ' Constellation' suffix if present and validate."""
    c = raw.strip()
    if c.endswith(" Constellation"):
        c = c[: -len(" Constellation")]
    if c not in _CONSTELLATIONS:
        raise ValueError(
            f"Unknown constellation {raw!r} - expected one of {sorted(_CONSTELLATIONS)}"
        )
    return c


def _load_unabridged_overlay(
    path: Path = UNABRIDGED_OVERLAY_JSON,
) -> tuple[list[dict], list[str]]:
    """Return ``(jump_patches, repeatable_perks)`` from the overlay file."""
    if not path.exists():
        return [], []
    data = json.loads(path.read_text())
    patches = list(data.get("jump_patches") or [])
    repeatable = list(data.get("repeatable_perks") or [])
    return patches, repeatable


def _apply_unabridged_overlay(
    parsed_rows: list[dict],
    overlay_patches: list[dict],
) -> None:
    """Mutate ``parsed_rows`` in place per the overlay patches.

    Each row is ``{constellation, jump, status, names: list[str]}``. A
    patch keyed by ``(constellation, jump)`` removes/adds names from
    every matching row.
    """
    for patch in overlay_patches:
        target_const = patch.get("constellation")
        target_jump = patch.get("jump")
        remove_names = patch.get("remove_names") or []
        add_names = patch.get("add_names") or []
        matching = [
            r for r in parsed_rows
            if r["constellation"] == target_const and r["jump"] == target_jump
        ]
        if not matching:
            raise ValueError(
                f"unabridged_overlay.json: no parsed Unabridged List row "
                f"matches (constellation={target_const!r}, "
                f"jump={target_jump!r}); curator must verify the patch key."
            )
        for row in matching:
            for nm in remove_names:
                row["names"] = [n for n in row["names"] if n != nm]
            # Add only if not already present anywhere in matching set,
            # to avoid duplicates when the patch targets multiple rows.
            existing = {n for r in matching for n in r["names"]}
            for nm in add_names:
                if nm not in existing:
                    matching[0]["names"].append(nm)
                    existing.add(nm)


def build_directory(
    wb,
    catalog_idx: dict,
    obtained_idx: dict,
    overlay_patches: list[dict] | None = None,
) -> tuple[list[DirectoryPerk], dict[str, int], int, set[tuple[str, str]]]:
    """Walk the Unabridged List sheet and build the directory.

    Returns (perks, status_distribution, expanded_excluded_count,
    matched_obtained_keys).
    """
    ws = wb["Unabridged List"]
    perks: list[DirectoryPerk] = []
    status_dist: Counter[str] = Counter()
    excluded_expanded = 0
    matched_obtained_keys: set[tuple[str, str]] = set()
    # Track ids we've already emitted so multi-perk rows that re-list
    # the same name (or accidental dupes between rows of the same
    # constellation+jump) don't double up. Threaded through to the
    # supplemental-pass too so we never emit duplicate ids.
    seen_ids: set[str] = set()

    # Pass 1: parse rows into a list of dicts so the overlay can patch
    # them before we build the directory entries.
    parsed_rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        const_raw = _norm_cell(ws.cell(r, 1).value)
        names_raw = _norm_cell(ws.cell(r, 2).value)
        jump = _norm_cell(ws.cell(r, 3).value)
        status = _norm_cell(ws.cell(r, 4).value)
        if not names_raw:
            continue

        names = [n.strip() for n in names_raw.split("\n") if n.strip()]

        if status == "Excluded":
            excluded_expanded += len(names)
            continue
        if status not in _KEEP_STATUSES:
            raise ValueError(
                f"Row {r}: unexpected status {status!r} (name={names_raw!r}, "
                f"jump={jump!r}). Update _KEEP_STATUSES if this is intentional."
            )

        if not const_raw:
            raise ValueError(f"Row {r}: missing constellation for name={names_raw!r}")
        constellation = _normalize_constellation(const_raw)

        if not jump:
            # All non-excluded rows have a jump in the current data;
            # keep this strict so we notice if that ever changes.
            raise ValueError(
                f"Row {r}: missing jump for {constellation}/{names_raw!r}"
            )

        parsed_rows.append({
            "constellation": constellation,
            "jump": jump,
            "status": status,
            "names": names,
        })

    # Pass 1.5: apply curator overlay patches (data/manual/unabridged_overlay.json).
    _apply_unabridged_overlay(parsed_rows, overlay_patches or [])

    # Pass 2: build directory entries from the (possibly patched) rows.
    for row_state in parsed_rows:
        constellation = row_state["constellation"]
        jump = row_state["jump"]
        status = row_state["status"]
        for name in row_state["names"]:
            pid = _id_for(constellation, jump, name)
            if pid in seen_ids:
                continue
            seen_ids.add(pid)

            cat = _catalog_lookup(catalog_idx, name, jump)
            obtained_matches = _obtained_lookup(obtained_idx, name, jump)

            cost: int | None = None
            cost_text: str | None = None
            cost_unit: str | None = None
            free = False
            repeatable = False
            description = ""

            if cat is not None:
                cost = cat.get("cost")
                cost_text = cat.get("cost_text")
                cost_unit = cat.get("cost_unit")
                free = bool(
                    cat.get("cost") == 0
                    and (cat.get("cost_text") or "").lower().startswith("free")
                )
                repeatable = bool(cat.get("repeatable"))
                description = cat.get("description", "") or ""
            elif obtained_matches:
                # Fall back to obtained_perks's cost info
                first = obtained_matches[0]
                cost = first.get("cost")
                cost_text = first.get("cost_text")
                cost_unit = first.get("cost_unit")
                free = bool(first.get("free"))
                description = first.get("perk_text", "") or ""

            acquired_chapter_num: str | None = None
            acquired_epub_sequence: int | None = None
            acquired_instances: list[dict] = []
            matched = False
            if obtained_matches:
                matched = True
                # Filter out free ride-along acquisitions: per the
                # gacha mechanic, free perks come bundled with a paid
                # parent and don't represent independent entries. The
                # paid parent already accounts for them. (When Joe
                # acquires a paid version of the same perk later, we
                # take that paid acquisition's chapter.)
                paid_matches = [m for m in obtained_matches if not m.get("free")]
                source_matches = paid_matches or obtained_matches
                earliest = _earliest(source_matches)
                acquired_chapter_num = earliest.get("chapter_num")
                acquired_epub_sequence = earliest.get("epub_sequence")
                # Sub-instances in chapter order, all (paid+free) so
                # the UI can show the full ride-along.
                for m in sorted(
                    obtained_matches,
                    key=lambda p: (
                        _norm_chapter(p.get("chapter_num", "0")),
                        p.get("epub_sequence", 0),
                    ),
                ):
                    acquired_instances.append({
                        "name": m.get("perk_name", ""),
                        "chapter_num": m.get("chapter_num", ""),
                        "epub_sequence": m.get("epub_sequence", 0),
                        "free": bool(m.get("free")),
                    })
                # Track every acquisition we covered so the spot-check
                # can report which obtained rows aren't represented in
                # the directory at all.
                for m in obtained_matches:
                    matched_obtained_keys.add(
                        (m.get("perk_name", ""), m.get("jump") or "")
                    )

            perks.append(
                DirectoryPerk(
                    id=pid,
                    constellation=constellation,
                    name=name,
                    jump=jump,
                    status=status,
                    cost=cost,
                    cost_text=cost_text,
                    cost_unit=cost_unit,
                    free=free,
                    repeatable=repeatable,
                    description=description,
                    acquired_chapter_num=acquired_chapter_num,
                    acquired_epub_sequence=acquired_epub_sequence,
                    first_acquired_at_word_offset=None,  # future work
                    matched_to_obtained=matched,
                    source="unabridged",
                    acquired_instances=acquired_instances,
                )
            )
            status_dist[status] += 1

    # Sort for deterministic output: by constellation, jump, then name.
    perks.sort(key=lambda p: (p.constellation, p.jump, p.name))
    return perks, dict(status_dist), excluded_expanded, matched_obtained_keys


def _supplemental_perks(
    obtained: dict,
    matched_keys: set[tuple[str, str]],
    catalog_idx: dict,
    seen_ids: set[str],
) -> tuple[list[DirectoryPerk], int, set[tuple[str, str]]]:
    """Build supplemental directory entries for paid acquisitions that
    don't map to any Unabridged List entry. Free ride-along perks are
    skipped (they come bundled with a paid parent).

    Returns (perks, free_orphan_count, raw_keys_now_covered) - the
    third element lists obtained (perk_name, jump) pairs in their
    ORIGINAL (un-aliased) form, for the spot-check.
    """
    # Group orphan acquisitions by (name, canonical_jump) to dedupe;
    # apply JUMP_ALIASES so a supplemental for "Civilian Equipment
    # Package"/"GUNNM/Battle Angel Alita" lands in the canonical "GUNNM"
    # constellation cluster, matching the rest of the directory.
    by_key: dict[tuple[str, str], list[dict]] = defaultdict(list)
    free_orphans = 0
    for p in obtained.get("perks", []):
        raw_jump = p.get("jump") or ""
        key = (p.get("perk_name", ""), raw_jump)
        if key in matched_keys:
            continue
        if p.get("free"):
            free_orphans += 1
            continue
        canon_jump = JUMP_ALIASES.get(raw_jump, raw_jump)
        by_key[(p.get("perk_name", ""), canon_jump)].append(p)

    out: list[DirectoryPerk] = []
    raw_keys_covered: set[tuple[str, str]] = set()
    for (name, jump), entries in sorted(by_key.items()):
        earliest = _earliest(entries)
        constellation = earliest.get("constellation") or ""
        if constellation not in _CONSTELLATIONS:
            # If obtained_perks left it blank, skip - we have nowhere
            # in the sky to put it. This shouldn't happen with current
            # data; raise to surface the case.
            raise ValueError(
                f"Supplemental perk {name!r}/{jump!r} has unknown "
                f"constellation {constellation!r}"
            )

        cat = _catalog_lookup(catalog_idx, name, jump)
        cost = earliest.get("cost")
        cost_text = earliest.get("cost_text")
        cost_unit = earliest.get("cost_unit")
        repeatable = False
        description = earliest.get("perk_text", "") or ""
        if cat is not None:
            cost = cat.get("cost", cost)
            cost_text = cat.get("cost_text", cost_text)
            cost_unit = cat.get("cost_unit", cost_unit)
            repeatable = bool(cat.get("repeatable"))
            if cat.get("description"):
                description = cat["description"]

        pid = _id_for(constellation, jump, name)
        if pid in seen_ids:
            continue  # already emitted (shouldn't happen, but be safe)
        seen_ids.add(pid)

        instances = []
        for m in sorted(
            entries,
            key=lambda p: (
                _norm_chapter(p.get("chapter_num", "0")),
                p.get("epub_sequence", 0),
            ),
        ):
            instances.append({
                "name": m.get("perk_name", ""),
                "chapter_num": m.get("chapter_num", ""),
                "epub_sequence": m.get("epub_sequence", 0),
                "free": bool(m.get("free")),
            })

        out.append(DirectoryPerk(
            id=pid,
            constellation=constellation,
            name=name,
            jump=jump,
            status="Obtained",
            cost=cost,
            cost_text=cost_text,
            cost_unit=cost_unit,
            free=False,
            repeatable=repeatable,
            description=description,
            acquired_chapter_num=earliest.get("chapter_num"),
            acquired_epub_sequence=earliest.get("epub_sequence"),
            first_acquired_at_word_offset=None,
            matched_to_obtained=True,
            source="obtained_supplemental",
            acquired_instances=instances,
        ))
        for m in entries:
            raw_keys_covered.add(
                (m.get("perk_name", ""), m.get("jump") or "")
            )
    return out, free_orphans, raw_keys_covered


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    catalog = json.loads(PERKS_CATALOG_JSON.read_text())
    obtained = json.loads(OBTAINED_JSON.read_text())

    aliases = load_perk_aliases()
    alias_lookup = build_alias_lookup(aliases)
    overlay_patches, overlay_repeatable = _load_unabridged_overlay()

    catalog_idx = _build_catalog_index(catalog)
    obtained_idx = _build_obtained_index(obtained, aliases, alias_lookup)

    perks, status_dist, excluded_expanded, matched_keys = build_directory(
        wb, catalog_idx, obtained_idx,
        overlay_patches=overlay_patches,
    )

    seen_ids = {p.id for p in perks}
    supplementals, free_orphans, supp_raw_keys = _supplemental_perks(
        obtained, matched_keys, catalog_idx, seen_ids
    )
    perks.extend(supplementals)
    matched_keys.update(supp_raw_keys)
    for _ in supplementals:
        status_dist["Obtained"] = status_dist.get("Obtained", 0) + 1

    perks.sort(key=lambda p: (p.constellation, p.jump, p.name))

    # Apply curator's explicit repeatable_perks overlay: any directory
    # row whose name appears in the overlay's repeatable_perks list is
    # flagged repeatable. This replaces the old heuristic that flipped
    # the bit on every name that happened to be a perk_aliases canonical
    # (which false-positived on Class A typo canonicals like
    # "Feel It Out"). The flag is now data-driven only.
    repeatable_overlay_set = set(overlay_repeatable)
    for p in perks:
        if p.name in repeatable_overlay_set:
            p.repeatable = True

    acquired = sum(1 for p in perks if p.matched_to_obtained)

    # Overlay assertions (run at end so they assert against the final
    # directory state, including supplementals): every add_name must
    # end up as a source="unabridged" directory row, every remove_name
    # must NOT appear under the patched (const, jump), and every
    # repeatable_perks entry must match at least one directory row.
    overlay_by_key: dict[tuple[str, str], dict] = {}
    for patch in overlay_patches:
        overlay_by_key[(patch["constellation"], patch["jump"])] = patch
    for (const_key, jump_key), patch in overlay_by_key.items():
        for added in patch.get("add_names") or []:
            ok = any(
                p.name == added
                and p.constellation == const_key
                and p.jump == jump_key
                and p.source == "unabridged"
                for p in perks
            )
            if not ok:
                raise AssertionError(
                    f"unabridged_overlay: add_name {added!r} for "
                    f"({const_key}, {jump_key}) is not a source='unabridged' "
                    f"directory row after build"
                )
        for removed in patch.get("remove_names") or []:
            still = any(
                p.name == removed
                and p.constellation == const_key
                and p.jump == jump_key
                for p in perks
            )
            if still:
                raise AssertionError(
                    f"unabridged_overlay: remove_name {removed!r} for "
                    f"({const_key}, {jump_key}) still appears in directory"
                )
    for name in overlay_repeatable:
        if not any(p.name == name for p in perks):
            raise AssertionError(
                f"unabridged_overlay: repeatable_perks entry {name!r} "
                f"matches no directory row"
            )

    write_validated_json(
        OUT,
        {
            "schema_version": 1,
            "_source": "data/raw/Brocktons_Celestial_Forge_Reference.xlsx#Unabridged List",
            "_count": len(perks),
            "_status_distribution": status_dist,
            "_acquired_count": acquired,
            "_supplemental_count": len(supplementals),
            "_free_ride_along_count": free_orphans,
            "_note": (
                "Master roster of every rollable perk in the gacha. Built from "
                "the Unabridged List (one row per perk; multi-perk rows in "
                "column B are expanded; 'Excluded' rows dropped) and "
                "supplemented by paid acquisitions from obtained_perks.json "
                "that have no Unabridged entry (source='obtained_supplemental', "
                "status='Obtained'). cost/cost_text/repeatable/description are "
                "joined from perks_catalog.json by (name, jump) with cosmetic- "
                "tolerant fallbacks. Acquisitions fold into the directory via "
                "exact, normalized, separator-prefix/suffix-split, and "
                "normalized-word-prefix matching, plus JUMP_ALIASES for "
                "cosmetic jump-name drift and data/manual/perk_aliases.json "
                "for canonical-name folding (typos, sub-instance variants). "
                "Free ride-along acquisitions (perks bundled with a paid parent) "
                "do NOT create supplemental entries - they're listed in the "
                "parent's acquired_instances."
            ),
            "perks": [asdict(p) for p in perks],
        },
        "perk_directory",
    )

    # ---- summary --------------------------------------------------------
    by_const = Counter(p.constellation for p in perks)
    print(f"wrote {OUT.relative_to(ROOT)}: {len(perks)} perks "
          f"(dropped {excluded_expanded} Excluded; "
          f"+{len(supplementals)} supplemental from obtained)")
    print(f"  acquired (matched to obtained_perks): {acquired}/{len(perks)}")
    print(f"  free ride-along acquisitions (in parents' instances): {free_orphans}")
    print("  status distribution:")
    for s in sorted(status_dist, key=lambda k: -status_dist[k]):
        print(f"    {s:12s}  {status_dist[s]}")
    print("  by constellation:")
    for c in sorted(by_const, key=lambda k: -by_const[k]):
        print(f"    {c:25s}  {by_const[c]}")

    # ---- spot check: after supplementals, every obtained acquisition
    # should map to a directory entry. Free ride-alongs are the
    # exception (they're in parent's acquired_instances, not their own
    # entry) - those are reported separately.
    obtained_keys = {
        (p.get("perk_name", ""), p.get("jump") or "")
        for p in obtained.get("perks", [])
    }
    paid_keys = {
        (p.get("perk_name", ""), p.get("jump") or "")
        for p in obtained.get("perks", []) if not p.get("free")
    }
    orphans = paid_keys - matched_keys
    print(f"  obtained_perks acquisitions: {obtained.get('_count', 0)} rows, "
          f"{len(obtained_keys)} unique (name, jump) pairs")
    print(f"  paid acquisitions matched:   "
          f"{len(paid_keys) - len(orphans)}/{len(paid_keys)}")
    if orphans:
        print(f"  PAID ORPHANS ({len(orphans)}) - paid acquisitions with no "
              "directory entry after supplemental pass:", file=sys.stderr)
        for name, jump in sorted(orphans):
            print(f"    - {name!r}  ({jump!r})", file=sys.stderr)


if __name__ == "__main__":
    main()
