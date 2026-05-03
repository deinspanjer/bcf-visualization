"""Parse the Comment Analysis sheet into comments.json.

Source:  data/raw/Brocktons_Celestial_Forge_Rolls_List.xlsx
         sheet "Comment Analysis"
Output:  data/derived/comments.json

The sheet has two parallel blocks (AO3 in cols 8-12, SV in cols 14-18),
each labelled like "12 - AO3" or "12 - SV". We merge them by chapter
prefix into a single per-chapter record. Each block has its own
"Last Checked" cell at the bottom (the AO3 and SV scrapes ran on
different dates), preserved as `_last_checked`.
"""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "raw" / "Brocktons_Celestial_Forge_Rolls_List.xlsx"
OUT = ROOT / "data" / "derived" / "comments.json"


@dataclass
class SiteStats:
    chapter_comments: int
    total_comments_so_far: int
    chapter_word_count: int
    total_word_count_so_far: int


@dataclass
class ChapterComments:
    chapter_num: str
    ao3: SiteStats | None = None
    sv: SiteStats | None = None


_LABEL_RE = re.compile(r"^\s*(?P<chap>\S+)\s*-\s*(?P<site>AO3|SV)\s*$")
_CHECKED_RE = re.compile(r"Last Checked:\s*(?P<date>.+)$", re.IGNORECASE)


def _read_block(ws, label_col: int) -> tuple[dict[str, SiteStats], str | None]:
    """Read one of the two side-by-side blocks (AO3 or SV).

    Returns (per-chapter stats keyed by chapter prefix, last-checked note).
    """
    stats: dict[str, SiteStats] = {}
    last_checked: str | None = None
    for r in range(2, ws.max_row + 1):
        label = ws.cell(r, label_col).value
        if not label:
            continue
        text = str(label).strip()
        m = _LABEL_RE.match(text)
        if not m:
            checked = _CHECKED_RE.search(text)
            if checked:
                last_checked = checked.group("date").strip()
            continue
        chap = m.group("chap")
        cells = [ws.cell(r, c).value for c in (label_col + 1, label_col + 2, label_col + 3, label_col + 4)]
        if any(v is None for v in cells):
            continue
        stats[chap] = SiteStats(
            chapter_comments=int(cells[0]),
            total_comments_so_far=int(cells[1]),
            chapter_word_count=int(cells[2]),
            total_word_count_so_far=int(cells[3]),
        )
    return stats, last_checked


def _chapter_sort_key(num: str) -> tuple[int, int]:
    parts = num.split(".", 1)
    try:
        major = int(parts[0])
        minor = int(parts[1]) if len(parts) > 1 else 0
    except ValueError:
        return (10**6, 0)
    return (major, minor)


def main() -> None:
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Comment Analysis"]

    ao3, ao3_checked = _read_block(ws, label_col=8)
    sv, sv_checked = _read_block(ws, label_col=14)

    chapters = sorted(set(ao3) | set(sv), key=_chapter_sort_key)
    rows = [
        ChapterComments(chapter_num=ch, ao3=ao3.get(ch), sv=sv.get(ch))
        for ch in chapters
    ]

    payload = {
        "_source": "data/raw/Brocktons_Celestial_Forge_Rolls_List.xlsx#Comment Analysis",
        "_count": len(rows),
        "_ao3_last_checked": ao3_checked,
        "_sv_last_checked": sv_checked,
        "comments": [
            {
                "chapter_num": r.chapter_num,
                "ao3": asdict(r.ao3) if r.ao3 else None,
                "sv": asdict(r.sv) if r.sv else None,
            }
            for r in rows
        ],
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")

    ao3_total = sum(r.ao3.total_comments_so_far for r in rows if r.ao3 and r.chapter_num == rows[-1].chapter_num)
    sv_total = sum(r.sv.total_comments_so_far for r in rows if r.sv and r.chapter_num == rows[-1].chapter_num)
    print(f"wrote {OUT.relative_to(ROOT)}: {len(rows)} chapters")
    print(f"  ao3 entries: {sum(1 for r in rows if r.ao3)}  last_checked={ao3_checked}")
    print(f"  sv  entries: {sum(1 for r in rows if r.sv)}  last_checked={sv_checked}")


if __name__ == "__main__":
    main()
